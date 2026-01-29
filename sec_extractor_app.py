"""
SEC EDGAR Financial Data Extractor - Streamlit Web Application
A modern web interface for extracting SEC financial data from multiple companies.
"""

import streamlit as st
import pandas as pd
import requests
import time
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Page configuration
st.set_page_config(
    page_title="SEC Financial Data Extractor",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main-header {
        font-size: 3rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f77b4;
        color: white;
        font-size: 1.1rem;
        padding: 0.75rem;
        border-radius: 8px;
    }
    .stButton>button:hover {
        background-color: #1557a0;
    }
    .success-box {
        padding: 1rem;
        border-radius: 8px;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
    }
    .info-box {
        padding: 1rem;
        border-radius: 8px;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
    }
    </style>
""", unsafe_allow_html=True)

# Company configurations
COMPANIES = {
    'Caterpillar Inc. (CAT)': {
        'name': 'Caterpillar Inc.',
        'ticker': 'CAT',
        'cik': '0000018230',
        'output_file': 'caterpillar_financials.xlsx'
    },
    'Deere & Company (DE)': {
        'name': 'Deere & Company',
        'ticker': 'DE',
        'cik': '0000315189',
        'output_file': 'deere_financials.xlsx'
    },
    'The Toro Company (TTC)': {
        'name': 'The Toro Company',
        'ticker': 'TTC',
        'cik': '0000097745',
        'output_file': 'toro_financials.xlsx'
    }
}


class SECEdgarExtractor:
    """Extract financial data from SEC EDGAR API"""
    
    def __init__(self, email, cik, company_name):
        self.base_url = "https://data.sec.gov"
        self.headers = {
            'User-Agent': f'{email}',
            'Accept-Encoding': 'gzip, deflate',
            'Host': 'data.sec.gov'
        }
        self.cik = cik
        self.company_name = company_name
        
    def get_company_facts(self):
        """Retrieve all company facts from SEC EDGAR API"""
        url = f"{self.base_url}/api/xbrl/companyfacts/CIK{self.cik}.json"
        
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            time.sleep(0.1)  # Rate limiting
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching data: {e}")
            raise
    
    def extract_financial_statement_data(self, facts_data, statement_type):
        """Extract specific financial statement data"""
        statement_items = self._get_statement_items(statement_type)
        us_gaap = facts_data.get('facts', {}).get('us-gaap', {})
        
        all_records = []
        
        for item_name, item_label in statement_items.items():
            if item_name in us_gaap:
                item_data = us_gaap[item_name]
                units = item_data.get('units', {})
                
                if 'USD' in units:
                    for record in units['USD']:
                        all_records.append({
                            'Line_Item': item_label,
                            'XBRL_Tag': item_name,
                            'Value': record.get('val'),
                            'End_Date': record.get('end'),
                            'Start_Date': record.get('start'),
                            'Filed_Date': record.get('filed'),
                            'Form': record.get('form'),
                            'Fiscal_Year': record.get('fy'),
                            'Fiscal_Period': record.get('fp'),
                            'Accession_Number': record.get('accn'),
                            'Frame': record.get('frame')
                        })
        
        if not all_records:
            return pd.DataFrame()
        
        df = pd.DataFrame(all_records)
        df['End_Date'] = pd.to_datetime(df['End_Date'])
        df['Filed_Date'] = pd.to_datetime(df['Filed_Date'])
        df['Start_Date'] = pd.to_datetime(df['Start_Date'])
        df = df.sort_values(['End_Date', 'Line_Item'], ascending=[False, True])
        
        return df
    
    def _get_statement_items(self, statement_type):
        """Get relevant line items for each statement type"""
        if statement_type == 'income':
            return {
                'Revenues': '     Total sales and revenues',
                'CostOfRevenue': '     Cost of goods sold',
                'SellingGeneralAndAdministrativeExpense': '     SG&A Expenses',
                'ResearchAndDevelopmentExpense': '     R&D Expenses',
                'FinancingInterestExpense': '     Interest expense of Financial Products',
                'OtherOperatingIncomeExpenseNet': '     Other operating (income) expenses',
                'CostsAndExpenses': '     Total operating costs',
                'OperatingIncomeLoss': 'Operating Profit',
                'InterestExpenseNonoperating': '     Interest expense excluding Financial Products',
                'OtherNonoperatingIncomeExpense': '     Other income (expense)',
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': 'Consolidated profit before taxes',
                'IncomeTaxExpenseBenefit': '     Provision (benefit) for income taxes',
                'IncomeLossFromEquityMethodInvestments': '     Equity in profit (loss) of unconsolidated affiliated companies',
                'ProfitLoss': 'Profit of consolidated and affiliated companies',
                'NetIncomeLossAttributableToNoncontrollingInterest': 'Less: Profit (loss) attributable to noncontrolling interests',
                'NetIncomeLossAvailableToCommonStockholdersBasic': 'Profit (Attributable to Common Stockholders)',
                'EarningsPerShareBasic': 'Profit per common share',
                'EarningsPerShareDiluted': 'Profit per common share - diluted',
                'WeightedAverageNumberOfSharesOutstandingBasic': 'Shares Outstanding - Basic',
                'WeightedAverageNumberOfDilutedSharesOutstanding': 'Shares Outstanding - Diluted',
            }
        
        elif statement_type == 'balance':
            return {
                'CashAndCashEquivalentsAtCarryingValue': '     Cash and cash equivalents',
                'ShortTermInvestments': '     Short-term investments',
                'AccountsReceivableNetCurrent': '     Receivables - trade and other',
                'InventoryNet': '     Inventories',
                'PrepaidExpenseAndOtherAssetsCurrent': '     Other current assets',
                'AssetsCurrent': 'Total current assets',
                'PropertyPlantAndEquipmentNet': '     Property, plant and equipment - net',
                'LongTermInvestments': '     Long-term investments',
                'Goodwill': '     Goodwill',
                'IntangibleAssetsNetExcludingGoodwill': '     Intangible assets',
                'OtherAssetsNoncurrent': '     Other assets',
                'Assets': 'Total Assets',
                'ShortTermBorrowings': '     Short-term borrowings',
                'AccountsPayableCurrent': '     Accounts payable',
                'AccruedLiabilitiesCurrent': '     Accrued expenses',
                'LongTermDebtCurrent': '     Current portion of long-term debt',
                'LiabilitiesCurrent': 'Total current liabilities',
                'LongTermDebtNoncurrent': '     Long-term debt',
                'DeferredTaxLiabilitiesNoncurrent': '     Deferred income taxes',
                'OtherLiabilitiesNoncurrent': '     Other liabilities',
                'Liabilities': 'Total Liabilities',
                'CommonStockValue': '     Common stock',
                'RetainedEarningsAccumulatedDeficit': '     Retained earnings',
                'AccumulatedOtherComprehensiveIncomeLossNetOfTax': '     Accumulated other comprehensive income',
                'StockholdersEquity': 'Total Stockholders\' Equity',
                'LiabilitiesAndStockholdersEquity': 'Total Liabilities and Stockholders\' Equity',
            }
        
        elif statement_type == 'cashflow':
            return {
                'ProfitLoss': 'Profit of consolidated and affiliated companies',
                'DepreciationDepletionAndAmortization': '     Depreciation and amortization',
                'IncreaseDecreaseInAccountsReceivable': '     Receivables',
                'IncreaseDecreaseInInventories': '     Inventories',
                'IncreaseDecreaseInAccountsPayable': '     Accounts payable and accrued expenses',
                'OtherOperatingActivitiesNet': '     Other operating activities - net',
                'NetCashProvidedByUsedInOperatingActivities': 'Cash provided by (used for) operating activities',
                'PaymentsToAcquirePropertyPlantAndEquipment': '     Capital expenditures',
                'PaymentsToAcquireBusinessesNetOfCashAcquired': '     Acquisitions',
                'ProceedsFromSaleOfPropertyPlantAndEquipment': '     Proceeds from disposals',
                'PaymentsToAcquireInvestments': '     Purchases of investments',
                'ProceedsFromSaleAndMaturityOfInvestments': '     Proceeds from investments',
                'NetCashProvidedByUsedInInvestingActivities': 'Cash provided by (used for) investing activities',
                'ProceedsFromIssuanceOfLongTermDebt': '     Proceeds from debt',
                'RepaymentsOfLongTermDebt': '     Payments on debt',
                'PaymentsOfDividends': '     Dividends paid',
                'PaymentsForRepurchaseOfCommonStock': '     Stock repurchases',
                'NetCashProvidedByUsedInFinancingActivities': 'Cash provided by (used for) financing activities',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect': 'Effect of exchange rate changes on cash',
                'CashAndCashEquivalentsPeriodIncreaseDecrease': 'Increase (decrease) in cash and cash equivalents',
            }
        
        return {}
    
    def create_pivot_table(self, df, statement_type):
        """Create pivot table view of the data"""
        if df.empty:
            return pd.DataFrame()
        
        annual_df = df[df['Form'].isin(['10-K', '10-Q'])].copy()
        
        if annual_df.empty:
            return df
        
        if statement_type == 'income':
            annual_df['Period_Length'] = (annual_df['End_Date'] - annual_df['Start_Date']).dt.days
            annual_df = annual_df[annual_df['Period_Length'] <= 100]
            if annual_df.empty:
                return pd.DataFrame()
        
        pivot = annual_df.pivot_table(
            index='Line_Item',
            columns='End_Date',
            values='Value',
            aggfunc='first'
        )
        
        pivot = pivot[sorted(pivot.columns, reverse=True)]
        pivot.columns = [col.strftime('%Y-%m-%d') if isinstance(col, datetime) else str(col)
                        for col in pivot.columns]
        
        desired_order = list(self._get_statement_items(statement_type).values())
        ordered_index = [lbl for lbl in desired_order if lbl in pivot.index]
        ordered_index += [lbl for lbl in pivot.index if lbl not in ordered_index]
        pivot = pivot.reindex(ordered_index)
        
        return pivot
    
    def format_excel_sheet(self, writer, sheet_name, df):
        """Apply formatting to Excel sheet"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        header_format = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.fill = header_format
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    text = str(cell.value) if cell.value is not None else ""
                    if len(text) > max_length:
                        max_length = len(text)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        accounting_format = '#,##0'
        for row in worksheet.iter_rows(min_row=2):
            for cell in row[1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = accounting_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        worksheet.freeze_panes = 'B2'
    
    def export_to_excel(self, output_filename, progress_callback=None):
        """Main function to extract all data and export to Excel"""
        
        if progress_callback:
            progress_callback(0.1, "Fetching company data from SEC EDGAR...")
        
        facts_data = self.get_company_facts()
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            if progress_callback:
                progress_callback(0.3, "Processing Income Statement...")
            
            income_df = self.extract_financial_statement_data(facts_data, 'income')
            if not income_df.empty:
                income_df.to_excel(writer, sheet_name='Income Statement - Raw', index=False)
                self.format_excel_sheet(writer, 'Income Statement - Raw', income_df)
                
                income_pivot = self.create_pivot_table(income_df, 'income')
                if not income_pivot.empty:
                    income_pivot.to_excel(writer, sheet_name='Income Statement - Quarterly')
                    self.format_excel_sheet(writer, 'Income Statement - Quarterly', income_pivot)
            
            if progress_callback:
                progress_callback(0.5, "Processing Balance Sheet...")
            
            balance_df = self.extract_financial_statement_data(facts_data, 'balance')
            if not balance_df.empty:
                balance_df.to_excel(writer, sheet_name='Balance Sheet - Raw', index=False)
                self.format_excel_sheet(writer, 'Balance Sheet - Raw', balance_df)
                
                balance_pivot = self.create_pivot_table(balance_df, 'balance')
                if not balance_pivot.empty:
                    balance_pivot.to_excel(writer, sheet_name='Balance Sheet - Quarterly')
                    self.format_excel_sheet(writer, 'Balance Sheet - Quarterly', balance_pivot)
            
            if progress_callback:
                progress_callback(0.8, "Processing Cash Flow Statement...")
            
            cashflow_df = self.extract_financial_statement_data(facts_data, 'cashflow')
            if not cashflow_df.empty:
                cashflow_df.to_excel(writer, sheet_name='Cash Flow - Raw', index=False)
                self.format_excel_sheet(writer, 'Cash Flow - Raw', cashflow_df)
                
                cashflow_pivot = self.create_pivot_table(cashflow_df, 'cashflow')
                if not cashflow_pivot.empty:
                    cashflow_pivot.to_excel(writer, sheet_name='Cash Flow - Quarterly')
                    self.format_excel_sheet(writer, 'Cash Flow - Quarterly', cashflow_pivot)
        
        if progress_callback:
            progress_callback(1.0, "Complete!")
        
        output.seek(0)
        return output


def main():
    # Header
    st.markdown('<div class="main-header">📊 SEC EDGAR Financial Data Extractor</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Extract complete financial statement history from SEC EDGAR database</div>', unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # Email input
        email = st.text_input(
            "Email Address",
            placeholder="your.email@company.com",
            help="Required by SEC API for identification"
        )
        
        st.markdown("---")
        
        # Company selection
        st.header("🏢 Select Company")
        selected_company = st.selectbox(
            "Choose a company:",
            options=list(COMPANIES.keys()),
            index=0
        )
        
        company_info = COMPANIES[selected_company]
        
        # Display company info
        st.info(f"""
        **Company:** {company_info['name']}  
        **Ticker:** {company_info['ticker']}  
        **CIK:** {company_info['cik']}
        """)
        
        st.markdown("---")
        
        # About section
        with st.expander("ℹ️ About"):
            st.markdown("""
            This application extracts financial data from the SEC EDGAR database.
            
            **Data Included:**
            - Income Statement
            - Balance Sheet
            - Cash Flow Statement
            
            **Formats:**
            - Raw data (all filings)
            - Quarterly pivot tables
            """)
    
    # Main content area
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        # Validation
        if not email or '@' not in email:
            st.warning("⚠️ Please enter a valid email address in the sidebar to continue.")
            st.stop()
        
        # Extract button
        if st.button("🚀 Extract Financial Data", use_container_width=True):
            
            # Progress tracking
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            try:
                # Create extractor
                extractor = SECEdgarExtractor(
                    email=email,
                    cik=company_info['cik'],
                    company_name=company_info['name']
                )
                
                # Define progress callback
                def update_progress(progress, message):
                    progress_bar.progress(progress)
                    status_text.text(message)
                
                # Extract data
                excel_data = extractor.export_to_excel(
                    output_filename=company_info['output_file'],
                    progress_callback=update_progress
                )
                
                # Success message
                st.success("✅ Financial data extracted successfully!")
                
                # Download button
                st.download_button(
                    label="📥 Download Excel File",
                    data=excel_data,
                    file_name=company_info['output_file'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                # Info box
                st.markdown(f"""
                <div class="info-box">
                <strong>📄 File Contents:</strong><br>
                • Income Statement (Raw + Quarterly)<br>
                • Balance Sheet (Raw + Quarterly)<br>
                • Cash Flow Statement (Raw + Quarterly)<br>
                <br>
                <strong>Company:</strong> {company_info['name']} ({company_info['ticker']})
                </div>
                """, unsafe_allow_html=True)
                
            except Exception as e:
                st.error(f"❌ Error: {str(e)}")
                st.info("Please check your internet connection and try again.")
    
    # Footer
    st.markdown("---")
    st.markdown(
        '<div style="text-align: center; color: #666; font-size: 0.9rem;">'
        'Data source: SEC EDGAR Database | Built with Streamlit'
        '</div>',
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
