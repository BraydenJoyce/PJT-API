import re
import time
import json
import logging
from datetime import datetime
from collections import OrderedDict
from typing import Optional, List, Dict

import requests
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Logging
# -----------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class ComprehensiveXBRLExtractor:
    """Extract complete financial statements with segment breakdowns"""

    def __init__(self, email, cik, company_name, ticker):
        self.base_url = "https://data.sec.gov"
        self.sec_archives = "https://www.sec.gov/Archives/edgar/data"
        self.headers = {
            # SEC fair access: identifiable User-Agent with email. Keep moderate request pacing.
            "User-Agent": f"{email}",
            "Accept-Encoding": "gzip, deflate",
        }
        self.cik = cik  # e.g., "0000018230"
        self.cik_int = str(int(cik))  # e.g., "18230"
        self.company_name = company_name
        self.ticker = ticker.lower()

        # XBRL namespaces
        self.namespaces = {
            "xbrli": "http://www.xbrl.org/2003/instance",
            "xbrldi": "http://xbrl.org/2006/xbrldi",
        }

    # -------------------------------------------------------------------------
    # Statement layouts
    # -------------------------------------------------------------------------
    def _get_statement_items(self, statement_type):
        if statement_type == "income":
            return OrderedDict([
                # Sales and Revenues
                ('Revenues_MET', ' Sales of Machinery, Energy & Transportation'),
                ('Revenues_FinancialProducts', ' Revenues of Financial Products'),
                ('Revenues_Total', ' Total sales and revenues'),

                # Operating Costs
                ('CostOfRevenue', ' Cost of goods sold'),
                ('SellingGeneralAndAdministrativeExpense', ' SG&A Expenses'),
                ('ResearchAndDevelopmentExpense', ' R&D Expenses'),
                ('FinancingInterestExpense_FinancialProducts', ' Interest expense of Financial Products'),
                ('OtherOperatingIncomeExpenseNet', ' Other operating (income) expenses'),
                ('CostsAndExpenses', ' Total operating costs'),

                ('OperatingIncomeLoss', 'Operating Profit'),
                ('InterestExpenseNonoperating_EXFP', ' Interest expense excluding Financial Products'),
                ('OtherNonoperatingIncomeExpense', ' Other income (expense)'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'Consolidated profit before taxes'),
                ('IncomeTaxExpenseBenefit', ' Provision (benefit) for income taxes'),
                ('ProfitOfConsolidatedCompanies', ' Profit of consolidated companies'),
                ('IncomeLossFromEquityMethodInvestments', ' Equity in profit (loss) of unconsolidated affiliated companies'),
                ('ProfitLoss', 'Profit of consolidated and affiliated companies'),
                ('NetIncomeLossAttributableToNoncontrollingInterest', 'Less: Profit (loss) attributable to noncontrolling interests'),
                ('NetIncomeLossAvailableToCommonStockholdersBasic', 'Profit (Attributable to Common Stockholders)'),

                # EPS
                ('EarningsPerShareBasic', 'Profit per common share'),
                ('EarningsPerShareDiluted', 'Profit per common share - diluted'),

                # Shares Outstanding
                ('WeightedAverageNumberOfSharesOutstandingBasic', 'Shares Outstanding - Basic'),
                ('WeightedAverageNumberOfDilutedSharesOutstanding', 'Shares Outstanding - Diluted'),
            ])

        elif statement_type == "balance":
            return OrderedDict([
                # Assets - Current
                ('CashAndCashEquivalentsAtCarryingValue', ' Cash & Cash Equivalents'),
                ('AccountsReceivableNetCurrent', ' Receivables - trade and other'),
                ('NotesAndLoansReceivableNetCurrent', ' Receivables - finance'),
                ('PrepaidExpenseAndOtherAssetsCurrent', ' Prepaid Expenses And Other Assets Current'),
                ('InventoryNet', ' Inventories'),
                ('AssetsCurrent', ' Total Current Assets'),

                # Assets - Noncurrent / Total
                ('PropertyPlantAndEquipmentNet', ' Property, Plant, & Equipment - net'),
                ('AccountsReceivableNetNoncurrent', ' Long-term receivables - trade and other'),
                ('NotesAndLoansReceivableNetNoncurrent', ' Long-term receivables - finance'),
                ('NoncurrentDeferredAndRefundableIncomeTaxes', ' Noncurrent deferred and refundable income taxes'),
                ('IntangibleAssetsNetExcludingGoodwill', ' Intangible Assets'),
                ('Goodwill', ' Goodwill'),
                ('OtherAssetsNoncurrent', ' Other assets'),
                ('Assets', 'Total assets'),

                # Liabilities - Current
                ('ShortTermBorrowings_FinancialProducts', ' Financial Products'),
                ('AccountsPayableCurrent', ' Accounts payable'),
                ('AccruedLiabilitiesCurrent', ' Accrued expenses'),
                ('EmployeeRelatedLiabilitiesCurrent', ' Accrued wages: salaries and employee benefits'),
                ('ContractWithCustomerLiabilityCurrent', ' Customer advances'),
                ('DividendsPayableCurrent', ' Dividends payable'),
                ('OtherLiabilitiesCurrent', ' Other current liabilities'),

                # Liabilities - Long-term / Total
                ('LongTermDebtAndCapitalLeaseObligationsCurrent_MET', ' Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligationsCurrent_FP', ' Financial Products'),
                ('LiabilitiesCurrent', ' Total current liabilities'),
                ('LongTermDebtAndCapitalLeaseObligations_MET', ' Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligations_FP', ' Financial Products'),
                ('PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent', ' Liability for postemployment benefits'),
                ('OtherLiabilitiesNoncurrent', ' Other liabilities'),
                ('Liabilities', 'Total Liabilities'),

                # Equity
                ('CommonStocksIncludingAdditionalPaidInCapital', ' Issued shares at paid-in amount'),
                ('TreasuryStockValue', ' Treasury stock at cost'),
                ('RetainedEarningsAccumulatedDeficit', ' Profit employed in the business'),
                ('AccumulatedOtherComprehensiveIncomeLossNetOfTax', ' Accumulated other comprehensive income (loss)'),
                ('MinorityInterest', ' Noncontrolling interests'),
                ('StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', "Total shareholders' equity"),
                ('LiabilitiesAndStockholdersEquity', "Total liabilities and shareholders' equity"),
            ])

        elif statement_type == "cashflow":
            return OrderedDict([
                # Operating Activities
                ('ProfitLoss', ' Profit of consolidated and affiliated companies'),

                # Adjustments
                ('DepreciationDepletionAndAmortization', ' Depreciation and Amortization'),
                ('DeferredIncomeTaxExpenseBenefit', ' Provision (benefit) for deferred income taxes'),
                ('NonCashGainLossOnDivestiture', ' (Gain) loss on divestiture'),
                ('OtherNoncashIncomeExpense', ' Other'),

                # Changes in assets and liabilities
                ('IncreaseDecreaseInReceivables', ' Receivables – trade and other'),
                ('IncreaseDecreaseInInventories', ' Inventories'),
                ('IncreaseDecreaseInAccountsPayable', ' Accounts payable'),
                ('IncreaseDecreaseInAccruedLiabilities', ' Accrued expenses'),
                ('IncreaseDecreaseInEmployeeRelatedLiabilities', ' Accrued wages, salaries and employee benefits'),
                ('IncreaseDecreaseInContractWithCustomerLiability', ' Customer advances'),
                ('IncreaseDecreaseInOtherOperatingAssets', ' Other assets - net'),
                ('IncreaseDecreaseInOtherOperatingLiabilities', ' Other liabilities - net'),

                ('NetCashProvidedByUsedInOperatingActivities', 'Net cash provided by (used for) operating activities'),

                # Investing Activities
                ('PaymentsToAcquirePropertyPlantAndEquipment', ' Capital expenditures – excluding equipment leased to others'),
                ('PaymentsToAcquireEquipmentOnLease', ' Expenditures for equipment leased to others'),
                ('ProceedsFromSaleOfPropertyPlantAndEquipment', ' Proceeds from disposals of leased assets and property, plant and equipment'),
                ('PaymentsToAcquireFinanceReceivables', ' Additions to finance receivables'),
                ('ProceedsFromCollectionOfFinanceReceivables', ' Collections of finance receivables'),
                ('ProceedsFromSaleOfFinanceReceivables', ' Proceeds from sale of finance receivables'),
                ('PaymentsToAcquireBusinessesNetOfCashAcquired', ' Investments and acquisitions (net of cash acquired)'),
                ('ProceedsFromDivestitureOfBusinessesNetOfCashDivested', ' Proceeds from sale of businesses and investments (net of cash sold)'),
                ('ProceedsFromSaleAndMaturityOfMarketableSecurities', ' Proceeds from maturities and sale of securities'),
                ('PaymentsToAcquireMarketableSecurities', ' Investments in securities'),
                ('PaymentsForProceedsFromOtherInvestingActivities', ' Other – net'),

                ('NetCashProvidedByUsedInInvestingActivities', 'Net cash provided by (used for) investing activities'),

                # Financing Activities
                ('PaymentsOfDividendsCommonStock', ' Dividends paid'),
                ('ProceedsFromIssuanceOrSaleOfEquity', ' Common stock issued, and other stock compensation transactions, net'),
                ('PaymentsForRepurchaseOfCommonStock', ' Payments to purchase common stock'),
                ('PaymentsForExciseTaxOnPurchaseOfCommonStock', ' Excise tax paid on purchases of common stock'),

                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_MET', ' - Machinery, Energy & Transportation'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_FP', ' - Financial Products'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_MET', ' - Machinery, Energy & Transportation'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_FP', ' - Financial Products'),

                ('ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess', ' Short-term borrowings – net (original maturities three months or less)'),

                ('NetCashProvidedByUsedInFinancingActivities', 'Net cash provided by (used for) financing activities'),

                ('EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents', 'Effect of exchange rate changes on cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect', 'Increase (decrease) in cash, cash equivalents and restricted cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning', 'Cash, cash equivalents and restricted cash at beginning of period'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End', 'Cash, cash equivalents and restricted cash at end of period'),
            ])

        logger.warning(f"Unknown statement type: {statement_type}")
        return OrderedDict()

    # Candidate tag catalogs
    def _get_income_tag_candidates(self) -> Dict[str, List[str]]:
        return {
            'Revenues': [
                'Revenues', 'SalesRevenueNet', 'SalesAndRevenue', 'SalesRevenueGoodsNet',
                'SalesRevenueServicesNet', 'RevenueFromContractWithCustomerExcludingAssessedTax'
            ],
            'CostOfRevenue': ['CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfSales'],
            'SellingGeneralAndAdministrativeExpense': ['SellingGeneralAndAdministrativeExpense'],
            'ResearchAndDevelopmentExpense': ['ResearchAndDevelopmentExpense'],
            'OtherOperatingIncomeExpenseNet': ['OtherOperatingIncomeExpenseNet', 'OtherOperatingIncomeExpense'],
            'CostsAndExpenses': ['CostsAndExpenses', 'OperatingExpenses'],
            'OperatingIncomeLoss': ['OperatingIncomeLoss'],
            'InterestExpenseNonoperating': ['InterestExpenseNonoperating', 'InterestAndDebtExpense'],
            'InterestExpenseNonoperating_EXFP': ['InterestExpenseExcludingFinancialProducts'],
            'OtherNonoperatingIncomeExpense': ['OtherNonoperatingIncomeExpense', 'NonoperatingIncomeExpense'],
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': [
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments',
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxes'
            ],
            'IncomeTaxExpenseBenefit': ['IncomeTaxExpenseBenefit'],
            'IncomeLossFromEquityMethodInvestments': ['IncomeLossFromEquityMethodInvestments'],
            'ProfitLoss': ['ProfitLoss', 'NetIncomeLoss'],
            'NetIncomeLossAttributableToNoncontrollingInterest': ['NetIncomeLossAttributableToNoncontrollingInterest'],
            'NetIncomeLossAvailableToCommonStockholdersBasic': ['NetIncomeLossAvailableToCommonStockholdersBasic'],
            'EarningsPerShareBasic': ['EarningsPerShareBasic'],
            'EarningsPerShareDiluted': ['EarningsPerShareDiluted'],
            'WeightedAverageNumberOfSharesOutstandingBasic': ['WeightedAverageNumberOfSharesOutstandingBasic'],
            'WeightedAverageNumberOfDilutedSharesOutstanding': ['WeightedAverageNumberOfDilutedSharesOutstanding'],
            'ProfitOfConsolidatedCompanies': ['IncomeLossFromContinuingOperationsAfterIncomeTaxes'],
            'FinancingInterestExpense_FinancialProducts': ['InterestExpense', 'InterestAndDebtExpense'],
        }

    def _get_balance_tag_candidates(self) -> Dict[str, List[str]]:
        return {
            'CashAndCashEquivalentsAtCarryingValue': [
                'CashAndCashEquivalentsAtCarryingValue', 'CashAndCashEquivalents', 'CashCashEquivalentsAndShortTermInvestments'
            ],
            'AccountsReceivableNetCurrent': [
                'AccountsReceivableNetCurrent', 'ReceivablesNetCurrent', 'TradeAccountsReceivableNetCurrent'
            ],
            'NotesAndLoansReceivableNetCurrent': [
                'NotesAndLoansReceivableNetCurrent', 'LoansReceivableNetCurrent', 'FinanceReceivablesNetCurrent'
            ],
            'PrepaidExpenseAndOtherAssetsCurrent': [
                'PrepaidExpenseAndOtherAssetsCurrent', 'PrepaidExpenseCurrent', 'PrepaidExpenseAndOtherCurrentAssets'
            ],
            'InventoryNet': ['InventoryNet', 'InventoriesNet', 'InventoryAndOtherCurrentAssets'],
            'AssetsCurrent': ['AssetsCurrent', 'CurrentAssets'],

            'PropertyPlantAndEquipmentNet': [
                'PropertyPlantAndEquipmentNet', 'PropertyPlantAndEquipmentExcludingLandNet'
            ],
            'AccountsReceivableNetNoncurrent': ['AccountsReceivableNetNoncurrent', 'ReceivablesNetNoncurrent'],
            'NotesAndLoansReceivableNetNoncurrent': [
                'NotesAndLoansReceivableNetNoncurrent', 'LoansReceivableNetNoncurrent', 'FinanceReceivablesNetNoncurrent'
            ],
            'NoncurrentDeferredAndRefundableIncomeTaxes': [
                'NoncurrentDeferredAndRefundableIncomeTaxes', 'DeferredTaxAssetsNetNoncurrent', 'DeferredIncomeTaxAssetsNetNoncurrent'
            ],
            'IntangibleAssetsNetExcludingGoodwill': ['IntangibleAssetsNetExcludingGoodwill', 'IntangibleAssetsNet'],
            'Goodwill': ['Goodwill'],
            'OtherAssetsNoncurrent': ['OtherAssetsNoncurrent'],
            'Assets': ['Assets', 'AssetsNet'],

            'ShortTermBorrowings': ['ShortTermBorrowings', 'CommercialPaper', 'ShortTermDebt'],
            'AccountsPayableCurrent': ['AccountsPayableCurrent'],
            'AccruedLiabilitiesCurrent': ['AccruedLiabilitiesCurrent', 'OtherAccruedLiabilitiesCurrent'],
            'EmployeeRelatedLiabilitiesCurrent': ['EmployeeRelatedLiabilitiesCurrent', 'EmployeeRelatedLiabilities'],
            'ContractWithCustomerLiabilityCurrent': ['ContractWithCustomerLiabilityCurrent', 'DeferredRevenueCurrent'],
            'DividendsPayableCurrent': ['DividendsPayableCurrent', 'DividendsPayable'],
            'OtherLiabilitiesCurrent': ['OtherLiabilitiesCurrent'],
            'LiabilitiesCurrent': ['LiabilitiesCurrent', 'CurrentLiabilities'],

            'LongTermDebtAndCapitalLeaseObligationsCurrent': [
                'LongTermDebtAndCapitalLeaseObligationsCurrent', 'LongTermDebtCurrent'
            ],
            'LongTermDebtAndCapitalLeaseObligations': [
                'LongTermDebtAndCapitalLeaseObligations', 'LongTermDebtNoncurrent', 'LongTermDebt'
            ],
            'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent': [
                'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent',
                'PensionAndOtherPostretirementDefinedBenefitPlansLiabilityNoncurrent'
            ],
            'OtherLiabilitiesNoncurrent': ['OtherLiabilitiesNoncurrent'],
            'Liabilities': ['Liabilities'],

            'CommonStocksIncludingAdditionalPaidInCapital': [
                'CommonStocksIncludingAdditionalPaidInCapital', 'CommonStockAndAdditionalPaidInCapital'
            ],
            'TreasuryStockValue': ['TreasuryStockValue', 'TreasuryStock'],
            'RetainedEarningsAccumulatedDeficit': ['RetainedEarningsAccumulatedDeficit', 'RetainedEarnings'],
            'AccumulatedOtherComprehensiveIncomeLossNetOfTax': [
                'AccumulatedOtherComprehensiveIncomeLossNetOfTax', 'AccumulatedOtherComprehensiveIncomeLoss'
            ],
            'MinorityInterest': ['MinorityInterest', 'NoncontrollingInterest'],
            'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest': [
                'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', 'StockholdersEquity'
            ],
            'LiabilitiesAndStockholdersEquity': [
                'LiabilitiesAndStockholdersEquity',
                'LiabilitiesAndStockholdersEquityIncludingPortionAttributableToNoncontrollingInterest'
            ],
        }

    def _get_cashflow_tag_candidates(self) -> Dict[str, List[str]]:
        # Expanded coverage & variants for CAT and similar filers
        return {
            # Net income used in reconciliation
            'ProfitLoss': [
                'ProfitLoss', 'NetIncomeLoss',
                'NetIncomeLossAvailableToCommonStockholdersBasic'  # fallback if only this is tagged in CF recon
            ],

            # Adjustments
            'DepreciationDepletionAndAmortization': [
                'DepreciationDepletionAndAmortization', 'DepreciationAndAmortization'
            ],
            'DeferredIncomeTaxExpenseBenefit': ['DeferredIncomeTaxExpenseBenefit', 'DeferredIncomeTaxExpense'],
            'NonCashGainLossOnDivestiture': [
                'NonCashGainLossOnDivestiture', 'GainLossOnDispositionOfBusinessesNet', 'GainLossOnSaleOfBusiness'
            ],
            'OtherNoncashIncomeExpense': ['OtherNoncashIncomeExpense', 'OtherNoncashItems'],

            # Working capital changes
            'IncreaseDecreaseInReceivables': [
                'IncreaseDecreaseInReceivables', 'IncreaseDecreaseInAccountsReceivable', 'IncreaseDecreaseInTradeAccountsReceivable'
            ],
            'IncreaseDecreaseInInventories': ['IncreaseDecreaseInInventories'],
            'IncreaseDecreaseInAccountsPayable': ['IncreaseDecreaseInAccountsPayable'],
            'IncreaseDecreaseInAccruedLiabilities': ['IncreaseDecreaseInAccruedLiabilities', 'IncreaseDecreaseInAccruedLiabilitiesCurrent'],
            'IncreaseDecreaseInEmployeeRelatedLiabilities': ['IncreaseDecreaseInEmployeeRelatedLiabilities'],
            'IncreaseDecreaseInContractWithCustomerLiability': ['IncreaseDecreaseInContractWithCustomerLiability', 'IncreaseDecreaseInDeferredRevenue'],
            'IncreaseDecreaseInOtherOperatingAssets': ['IncreaseDecreaseInOtherOperatingAssets', 'IncreaseDecreaseInOtherAssets'],
            'IncreaseDecreaseInOtherOperatingLiabilities': ['IncreaseDecreaseInOtherOperatingLiabilities', 'IncreaseDecreaseInOtherLiabilities'],

            # Operating CF total (add more variants)
            'NetCashProvidedByUsedInOperatingActivities': [
                'NetCashProvidedByUsedInOperatingActivities',
                'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations',
                'NetCashProvidedByUsedInOperatingActivitiesExcludingDiscontinuedOperations',
                'NetCashProvidedByUsedInOperatingActivitiesIncludingDiscontinuedOperations',
                'NetCashProvidedByUsedInOperatingActivitiesIndirectMethod'
            ],

            # Investing CF lines and total
            'PaymentsToAcquirePropertyPlantAndEquipment': [
                'PaymentsToAcquirePropertyPlantAndEquipment', 'PaymentsToAcquireProductiveAssets',
                'PaymentsForProceedsFromPropertyPlantAndEquipment'
            ],
            'PaymentsToAcquireEquipmentOnLease': ['PaymentsToAcquireEquipmentOnLease', 'PaymentsToAcquireEquipmentLeasedToOthers'],
            'ProceedsFromSaleOfPropertyPlantAndEquipment': ['ProceedsFromSaleOfPropertyPlantAndEquipment', 'ProceedsFromSaleOfProductiveAssets'],
            'PaymentsToAcquireFinanceReceivables': ['PaymentsToAcquireFinanceReceivables'],
            'ProceedsFromCollectionOfFinanceReceivables': ['ProceedsFromCollectionOfFinanceReceivables'],
            'ProceedsFromSaleOfFinanceReceivables': ['ProceedsFromSaleOfFinanceReceivables'],
            'PaymentsToAcquireBusinessesNetOfCashAcquired': [
                'PaymentsToAcquireBusinessesNetOfCashAcquired', 'PaymentsToAcquireBusinessesAndIntangiblesNetOfCashAcquired'
            ],
            'ProceedsFromDivestitureOfBusinessesNetOfCashDivested': [
                'ProceedsFromDivestitureOfBusinessesNetOfCashDivested', 'ProceedsFromSaleOfBusinessesNetOfCashDivested'
            ],
            'ProceedsFromSaleAndMaturityOfMarketableSecurities': [
                'ProceedsFromSaleAndMaturityOfMarketableSecurities',
                'ProceedsFromMaturitiesOfAvailableForSaleSecurities',
                'ProceedsFromSaleOfAvailableForSaleSecurities'
            ],
            'PaymentsToAcquireMarketableSecurities': [
                'PaymentsToAcquireMarketableSecurities', 'PaymentsToAcquireAvailableForSaleSecurities'
            ],
            'PaymentsForProceedsFromOtherInvestingActivities': [
                'PaymentsForProceedsFromOtherInvestingActivities', 'NetCashProvidedByUsedInOtherInvestingActivities'
            ],
            'NetCashProvidedByUsedInInvestingActivities': [
                'NetCashProvidedByUsedInInvestingActivities',
                'NetCashProvidedByUsedInInvestingActivitiesContinuingOperations',
                'NetCashProvidedByUsedInInvestingActivitiesIncludingDiscontinuedOperations'
            ],

            # Financing CF lines and total
            'PaymentsOfDividendsCommonStock': ['PaymentsOfDividendsCommonStock', 'PaymentsOfDividends'],
            'ProceedsFromIssuanceOrSaleOfEquity': [
                'ProceedsFromIssuanceOrSaleOfEquity', 'ProceedsFromIssuanceOfCommonStock', 'ProceedsFromShareBasedCompensationArrangements'
            ],
            'PaymentsForRepurchaseOfCommonStock': ['PaymentsForRepurchaseOfCommonStock', 'PaymentsForRepurchaseOfEquity'],
            'PaymentsForExciseTaxOnPurchaseOfCommonStock': ['PaymentsForExciseTaxOnPurchaseOfCommonStock', 'ExciseTaxOnShareRepurchasesPaid'],

            'ProceedsFromDebtMaturingInMoreThanThreeMonths': [
                'ProceedsFromDebtMaturingInMoreThanThreeMonths', 'ProceedsFromIssuanceOfLongTermDebt'
            ],
            'RepaymentsOfDebtMaturingInMoreThanThreeMonths': [
                'RepaymentsOfDebtMaturingInMoreThanThreeMonths', 'RepaymentsOfLongTermDebt'
            ],
            'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess': [
                'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess',
                'ProceedsFromRepaymentsOfShortTermDebt', 'NetBorrowingsUnderLineOfCredit'
            ],
            'NetCashProvidedByUsedInFinancingActivities': [
                'NetCashProvidedByUsedInFinancingActivities',
                'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations',
                'NetCashProvidedByUsedInFinancingActivitiesIncludingDiscontinuedOperations'
            ],

            # FX and cash bridge
            'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': [
                'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents',
                'EffectOfExchangeRateOnCashAndCashEquivalents'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect',
                'CashAndCashEquivalentsPeriodIncreaseDecrease'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsAtCarryingValueBeginningBalance',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsBeginningBalance',
                'CashAndCashEquivalentsAtCarryingValueBeginningBalance'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsAtCarryingValue',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents',
                'CashAndCashEquivalentsAtCarryingValue'
            ],
        }

    # -------------------------------------------------------------------------
    # EDGAR retrieval
    # -------------------------------------------------------------------------
    def get_all_filings(self, start_year=2020):
        url = f"{self.base_url}/submissions/CIK{self.cik}.json"
        logger.info(f"Fetching all filings since {start_year} for {self.company_name}")
        resp = requests.get(url, headers=self.headers)
        resp.raise_for_status()
        time.sleep(0.2)  # be gentle
        data = resp.json()
        recent = data["filings"]["recent"]

        filings = []
        for i in range(len(recent["form"])):
            filing_date = recent["filingDate"][i]
            filing_year = int(filing_date.split("-")[0])
            if filing_year >= start_year:
                form = recent["form"][i]
                if form in ["10-Q", "10-K"]:
                    filings.append({
                        "accession": recent["accessionNumber"][i],
                        "filing_date": filing_date,
                        "report_date": recent["reportDate"][i],
                        "form": form,
                        "primary_document": recent["primaryDocument"][i],
                    })

        filings.sort(key=lambda x: x["report_date"])
        logger.info(f"Found {len(filings)} filings since {start_year}")
        return filings

    def _filing_base_dir(self, accession: str) -> str:
        """Return folder URL for a filing (no trailing slash)."""
        accession_no_dash = accession.replace("-", "")
        return f"{self.sec_archives}/{self.cik_int}/{accession_no_dash}"

    def get_filing_items(self, accession: str) -> List[Dict[str, str]]:
        """
        Return directory items (from index.json if available; else parse HTML directory).
        """
        base_dir = self._filing_base_dir(accession)

        # Try index.json first (preferred)
        idx_url = f"{base_dir}/index.json"
        try:
            r = requests.get(idx_url, headers=self.headers)
            if r.status_code == 200:
                time.sleep(0.2)
                j = r.json()
                items = j.get("directory", {}).get("item", [])
                return [{"name": it.get("name", ""), "type": it.get("type", "")} for it in items]
        except Exception:
            pass  # fall through to HTML parse

        # Fallback: parse HTML directory listing for hrefs
        try:
            r = requests.get(base_dir + "/", headers=self.headers)
            r.raise_for_status()
            time.sleep(0.2)
            html = r.text
            hrefs = re.findall(r'href="([^"]+)"', html, flags=re.IGNORECASE)
            names = [h for h in hrefs if not h.startswith("http") and not h.startswith("?") and "/" not in h.strip("/")]
            names = list(dict.fromkeys(names))  # de-dup preserve order
            return [{"name": n, "type": ""} for n in names]
        except Exception as e:
            logger.error(f"Unable to list directory for accession {accession}: {e}")
            return []

    def pick_instance_from_items(self, items: List[Dict[str, str]]) -> Optional[str]:
        """
        Choose the XBRL instance document (EX-101.INS) from directory items.
        """
        names = [i.get("name") for i in items if i.get("name")]
        if not names:
            return None

        # Candidate XMLs
        xmls = [n for n in names if n.lower().endswith(".xml")]

        # Exclude non-instance files (linkbases, schemas, helpers)
        EXCLUDE_SUBSTR = ["_cal.xml", "_def.xml", "_lab.xml", "_pre.xml",
                          ".xsd", "filingsummary", "metalink", "schema"]
        xmls = [n for n in xmls if all(ex not in n.lower() for ex in EXCLUDE_SUBSTR)]

        # Prefer typical instance patterns first
        preferences = [
            "_htm.xml",            # frequent iXBRL pack pattern
            f"{self.ticker}-",     # contains ticker
        ]
        for pref in preferences:
            filtered = [n for n in xmls if pref in n.lower()]
            if filtered:
                return sorted(filtered, key=len)[0]

        # Otherwise, pick the first remaining XML (shortest name heuristic)
        if xmls:
            return sorted(xmls, key=len)[0]

        return None

    def download_file(self, url: str) -> bytes:
        """GET with simple pacing and error surface."""
        r = requests.get(url, headers=self.headers)
        r.raise_for_status()
        time.sleep(0.2)
        return r.content

    # -------------------------------------------------------------------------
    # XBRL parsing
    # -------------------------------------------------------------------------
    def parse_context_elements(self, root):
        contexts = {}
        for context in root.findall(".//xbrli:context", self.namespaces):
            context_id = context.get("id")
            period = context.find("xbrli:period", self.namespaces)
            instant = period.find("xbrli:instant", self.namespaces)
            start = period.find("xbrli:startDate", self.namespaces)
            end = period.find("xbrli:endDate", self.namespaces)
            info = {
                "id": context_id,
                "segments": {},
                "instant": instant.text if instant is not None else None,
                "start": start.text if start is not None else None,
                "end": end.text if end is not None else None,
            }
            entity = context.find("xbrli:entity", self.namespaces)
            if entity is not None:
                segment = entity.find("xbrli:segment", self.namespaces)
                if segment is not None:
                    for member in segment.findall(".//xbrldi:explicitMember", self.namespaces):
                        dim = member.get("dimension")
                        val = member.text
                        if ":" in val:
                            val = val.split(":")[1]
                        info["segments"][dim] = val
            contexts[context_id] = info
        return contexts

    def extract_facts_from_xbrl(self, xml_content: bytes):
        root = ET.fromstring(xml_content)

        # Capture/merge document namespaces
        for prefix, uri in root.attrib.items():
            if prefix.startswith("{http://www.w3.org/2000/xmlns/}"):
                ns_prefix = prefix.split("}")[1]
                self.namespaces[ns_prefix] = uri

        contexts = self.parse_context_elements(root)

        facts = []
        for elem in root.iter():
            tag_name = elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag
            context_ref = elem.get("contextRef")
            unit_ref = elem.get("unitRef")
            decimals = elem.get("decimals")

            if context_ref in contexts and elem.text:
                context = contexts[context_ref]

                # Assume consolidated unless a segment is present
                segment_name = "Consolidated"
                segment_dimension = None
                if context["segments"]:
                    for dim, member in context["segments"].items():
                        segment_name = member
                        segment_dimension = dim
                        break

                try:
                    value = float(elem.text)
                except (ValueError, TypeError):
                    continue

                facts.append({
                    "tag": tag_name,
                    "value": value,
                    "context_id": context_ref,
                    "segment": segment_name,
                    "dimension": segment_dimension,
                    "start_date": context["start"],
                    "end_date": context["end"],
                    "instant_date": context["instant"],
                    "decimals": decimals,
                    "unit": unit_ref,
                })

        return facts

    def process_filing(self, filing: dict):
        """
        Load the filing folder, locate the actual XBRL instance file, download, parse facts.
        """
        try:
            base_dir = self._filing_base_dir(filing["accession"])
            logger.info(f"Processing {filing['form']} from {filing['report_date']}")

            # List directory items
            items = self.get_filing_items(filing["accession"])
            instance_name = self.pick_instance_from_items(items)

            # Fallback: scan the primary document HTML for a .xml instance href
            if not instance_name and filing.get("primary_document"):
                primary_url = f"{base_dir}/{filing['primary_document']}"
                try:
                    html = self.download_file(primary_url).decode("utf-8", errors="ignore")
                    links = re.findall(r'href="([^"]+\\.xml)"', html, flags=re.IGNORECASE)
                    links = [l for l in links if not any(s in l.lower() for s in
                                                         ["_cal.xml", "_def.xml", "_lab.xml", "_pre.xml", ".xsd",
                                                          "filingsummary", "metalink", "schema"])]
                    if links:
                        instance_name = links[0].split("/")[-1]
                except Exception:
                    pass

            if not instance_name:
                logger.warning(f"Could not locate XBRL instance for accession {filing['accession']} — skipping")
                return []

            instance_url = f"{base_dir}/{instance_name}"
            xml_content = self.download_file(instance_url)
            facts = self.extract_facts_from_xbrl(xml_content)

            # annotate facts
            for fact in facts:
                fact["accession"] = filing["accession"]
                fact["filing_date"] = filing["filing_date"]
                fact["report_date"] = filing["report_date"]
                fact["form"] = filing["form"]

            logger.info(f" Extracted {len(facts)} facts from {instance_name}")
            return facts

        except requests.HTTPError as e:
            logger.error(f"HTTP error for filing {filing.get('accession')}: {e}")
            return []
        except Exception as e:
            logger.error(f"Error processing filing: {e}")
            return []

    def extract_all_data(self, start_year=2020):
        filings = self.get_all_filings(start_year=start_year)
        all_facts = []

        for i, filing in enumerate(filings, 1):
            logger.info(f"\n[{i}/{len(filings)}] " + "=" * 50)
            facts = self.process_filing(filing)
            all_facts.extend(facts)

        df = pd.DataFrame(all_facts)
        if not df.empty:
            for date_col in ["start_date", "end_date", "instant_date", "filing_date", "report_date"]:
                if date_col in df.columns:
                    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

            sort_col = "end_date" if "end_date" in df.columns else "instant_date"
            df = df.sort_values([sort_col, "tag", "segment"], ascending=[True, True, True])
            logger.info(f"\nTotal facts extracted: {len(df)}")
        return df

    # -------------------------------------------------------------------------
    # Quarter math (YTD-aware) + discrete CF normalization
    # -------------------------------------------------------------------------
    def _is_calendar_ytd(self, row):
        s = row.get("start_date")
        e = row.get("end_date")
        if pd.isna(s) or pd.isna(e):
            return False
        return s.month == 1 and s.day == 1 and s.year == e.year  # calendar FY; adjust if needed

    def calculate_q4_data(self, df):
        logger.info("\n" + "=" * 60)
        logger.info("Calculating Q4 data")
        logger.info("=" * 60)

        if df.empty:
            return df

        quarterly_df = df[df["form"] == "10-Q"].copy()
        annual_df = df[df["form"] == "10-K"].copy()

        q4_records = []
        unique_tags = df["tag"].dropna().unique()
        unique_segments = df["segment"].dropna().unique().tolist()
        if "Consolidated" not in unique_segments:
            unique_segments.append("Consolidated")

        for tag in unique_tags:
            for segment in unique_segments:
                annual_subset = annual_df[
                    (annual_df["tag"] == tag) &
                    ((annual_df["segment"] == segment) |
                     (segment == "Consolidated" and annual_df["segment"].isna()))
                ].copy()

                for _, annual_row in annual_subset.iterrows():
                    is_balance_sheet = pd.notna(annual_row.get("instant_date"))

                    if is_balance_sheet:
                        fiscal_year_end = annual_row["instant_date"]
                        if pd.isna(fiscal_year_end):
                            continue
                        q4_records.append({
                            "tag": tag,
                            "value": annual_row["value"],
                            "segment": segment,
                            "start_date": None,
                            "end_date": None,
                            "instant_date": fiscal_year_end,
                            "report_date": fiscal_year_end,
                            "form": "10-K (Q4)",
                            "accession": annual_row["accession"],
                            "filing_date": annual_row["filing_date"],
                            "context_id": f"Q4_{fiscal_year_end.year}_{segment}",
                            "dimension": annual_row.get("dimension"),
                            "decimals": annual_row.get("decimals"),
                            "unit": annual_row.get("unit"),
                        })
                        continue

                    fiscal_year_end = annual_row["end_date"]
                    if pd.isna(fiscal_year_end):
                        continue

                    fiscal_year = fiscal_year_end.year
                    q_subset = quarterly_df[
                        (quarterly_df["tag"] == tag) &
                        ((quarterly_df["segment"] == segment) |
                         (segment == "Consolidated" and quarterly_df["segment"].isna())) &
                        (quarterly_df["end_date"] > pd.Timestamp(year=fiscal_year - 1, month=12, day=31)) &
                        (quarterly_df["end_date"] <= fiscal_year_end)
                    ].copy()

                    if q_subset.empty:
                        continue

                    q_subset["is_ytd"] = q_subset.apply(self._is_calendar_ytd, axis=1)

                    annual_total = annual_row["value"]
                    q3_end = q_subset["end_date"].max()

                    if q_subset["is_ytd"].any():
                        latest_ytd = q_subset.sort_values("end_date").iloc[-1]
                        q4_value = annual_total - latest_ytd["value"]
                    else:
                        sum_quarters = q_subset["value"].sum()
                        q4_value = annual_total - sum_quarters
                        if len(q_subset) != 3:
                            logger.warning(
                                f"Discrete quarterly data incomplete for {tag} ({segment}) FY{fiscal_year}: "
                                f"have {len(q_subset)} quarters; Q4 computed as Annual - sum(available)"
                            )

                    q4_records.append({
                        "tag": tag,
                        "value": q4_value,
                        "segment": segment,
                        "start_date": q3_end + pd.Timedelta(days=1) if pd.notna(q3_end) else None,
                        "end_date": fiscal_year_end,
                        "instant_date": None,
                        "report_date": fiscal_year_end,
                        "form": "10-Q (Q4 Calculated)",
                        "accession": annual_row["accession"],
                        "filing_date": annual_row["filing_date"],
                        "context_id": f"Q4_{fiscal_year}_{segment}",
                        "dimension": annual_row.get("dimension"),
                        "decimals": annual_row.get("decimals"),
                        "unit": annual_row.get("unit"),
                    })

        if q4_records:
            q4_df = pd.DataFrame(q4_records)
            combined = pd.concat([df, q4_df], ignore_index=True)
            if "end_date" in combined.columns:
                combined = combined.sort_values(["end_date", "instant_date", "tag", "segment"])
            logger.info(f"Added {len(q4_records)} Q4 records")
            return combined
        return df

    # ========================================================================
    # FIXED: Simplified cash flow normalization - assumes all 10-Q data is YTD
    # ========================================================================
    def _normalize_quarters_to_discrete(self, df_quarters: pd.DataFrame) -> pd.DataFrame:
        """
        Convert YTD cash flow values in 10-Q filings to discrete quarterly values.
        
        Cash flow statements in 10-Q filings are almost always reported as YTD
        (year-to-date), meaning Q2 shows Jan-Jun total, Q3 shows Jan-Sep total, etc.
        This function converts those cumulative values into discrete quarterly values.
        
        Example: If YTD values are Q1=$100M, Q2=$250M, Q3=$400M
                 Discrete values become Q1=$100M, Q2=$150M, Q3=$150M
        """
        if df_quarters.empty:
            return df_quarters

        q = df_quarters.copy()
        q = q[q["form"].str.contains("10-Q", na=False)]
        q = q[q["end_date"].notna()].copy()

        # Normalize segment and derive grouping keys
        q["segment"] = q["segment"].fillna("Consolidated")
        q["fiscal_year"] = q["end_date"].dt.year

        q = q.sort_values(["tag", "segment", "fiscal_year", "end_date"])

        def _to_discrete(g):
            """
            Convert YTD cumulative values to discrete quarterly values.
            
            For cash flow in 10-Qs, we assume all values are YTD and use
            differencing to get discrete quarters. The .diff() returns NaN
            for Q1, which fillna() replaces with the original Q1 value.
            """
            if not g.empty and len(g) > 1:
                g = g.sort_values("end_date").copy()
                # diff() subtracts previous row from current row
                # fillna() keeps Q1 as-is since it has no previous row
                g["value"] = g["value"].diff().fillna(g["value"])
            return g

        q = q.groupby(["tag", "segment", "fiscal_year"], group_keys=False).apply(_to_discrete)
        return q

    # -------------------------------------------------------------------------
    # Pivoting
    # -------------------------------------------------------------------------
    def create_statement_pivot(self, df, statement_type):
        if df.empty:
            return pd.DataFrame()

        statement_items = self._get_statement_items(statement_type)

        if statement_type == "balance":
            date_col = "instant_date"
            df_filtered = df[df["instant_date"].notna()].copy()
        else:
            date_col = "end_date"
            df_filtered = df[df["end_date"].notna()].copy()

        if df_filtered.empty:
            return pd.DataFrame()

        # Quarterly only (include Q4 Calculated for flows)
        df_filtered = df_filtered[df_filtered["form"].str.contains("10-Q", na=False)].copy()

        # =====================================================================
        # CASH FLOW NORMALIZATION: Convert YTD to discrete for Q1/Q2/Q3
        # =====================================================================
        if statement_type == "cashflow":
            is_q4_calc = df_filtered["form"].str.contains("Q4 Calculated", na=False)
            q4_calc = df_filtered[is_q4_calc].copy()
            q10 = df_filtered[~is_q4_calc].copy()
            
            # Convert YTD cash flow data to discrete quarterly values
            q10 = self._normalize_quarters_to_discrete(q10)
            
            df_filtered = pd.concat([q10, q4_calc], ignore_index=True)

        df_filtered["segment"] = df_filtered["segment"].fillna("Consolidated")

        # Candidate tag maps
        income_map = self._get_income_tag_candidates() if statement_type == "income" else {}
        balance_map = self._get_balance_tag_candidates() if statement_type == "balance" else {}
        cash_map = self._get_cashflow_tag_candidates() if statement_type == "cashflow" else {}

        # Segment suffix mapping
        segment_map = {
            "FinancialProducts": "FinancialProductsMember",
            "FP": "FinancialProductsMember",
            "MET": "MachineryEnergyTransportationMember",
            "EXFP": "AllOtherExcludingFinancialProductsMember",
            "Total": "Consolidated",
        }

        pivot_data = []

        for tag_key, label in statement_items.items():
            if tag_key == "":
                pivot_data.append({"Line_Item": label})
                continue

            parts = tag_key.split("_", 1)
            base_key = parts[0]
            segment_suffix = parts[1] if len(parts) > 1 else None

            if statement_type == "income":
                candidate_tags = income_map.get(base_key, [base_key])
            elif statement_type == "balance":
                candidate_tags = balance_map.get(base_key, [base_key if segment_suffix else tag_key])
            elif statement_type == "cashflow":
                candidate_tags = cash_map.get(base_key, [base_key if segment_suffix else tag_key])
            else:
                candidate_tags = [tag_key]

            selected_subset = pd.DataFrame()

            if segment_suffix:
                target_segment = segment_map.get(segment_suffix, segment_suffix)
                # Prioritized candidate resolution for explicit segment
                for cand in candidate_tags:
                    sub = df_filtered[(df_filtered["tag"] == cand) & (df_filtered["segment"] == target_segment)]
                    if not sub.empty:
                        selected_subset = sub
                        break
                # Fallback: case-insensitive contains
                if selected_subset.empty:
                    for cand in candidate_tags:
                        sub = df_filtered[
                            (df_filtered["tag"] == cand) &
                            (df_filtered["segment"].str.contains(target_segment, case=False, na=False))
                        ]
                        if not sub.empty:
                            selected_subset = sub
                            break
            else:
                # Consolidated / no explicit segment requested
                for cand in candidate_tags:
                    sub = df_filtered[
                        (df_filtered["tag"] == cand) &
                        (df_filtered["segment"].isin(["Consolidated", ""]))
                    ]
                    if not sub.empty:
                        selected_subset = sub
                        break

                # Fallback: if still empty, accept any segment (some filers omit explicit consolidated member)
                if selected_subset.empty:
                    for cand in candidate_tags:
                        sub = df_filtered[(df_filtered["tag"] == cand)]
                        if not sub.empty:
                            selected_subset = sub
                            break

            # --- Additional safety net for CASH FLOW TOTALS only ---
            is_cf_total = (statement_type == "cashflow") and (base_key in {
                'NetCashProvidedByUsedInOperatingActivities',
                'NetCashProvidedByUsedInInvestingActivities',
                'NetCashProvidedByUsedInFinancingActivities'
            })
            if selected_subset.empty and is_cf_total:
                pattern = base_key.lower()
                sub = df_filtered[df_filtered['tag'].str.lower().str.contains(pattern, na=False)]
                if not sub.empty:
                    sub_pref = sub[sub['segment'].isin(['Consolidated', ''])]
                    selected_subset = sub_pref if not sub_pref.empty else sub
            # -------------------------------------------------------

            if not selected_subset.empty:
                pv = selected_subset.pivot_table(index="tag", columns=date_col, values="value", aggfunc="first")
                if not pv.empty:
                    row = {"Line_Item": label}
                    row.update(pv.iloc[0].to_dict())
                    pivot_data.append(row)

        if not pivot_data:
            return pd.DataFrame()

        result_df = pd.DataFrame(pivot_data)

        # Sort columns by date (most recent first), keep Line_Item first
        date_cols = [c for c in result_df.columns if c != "Line_Item"]
        date_cols_sorted = sorted(date_cols, reverse=True)
        result_df = result_df[["Line_Item"] + date_cols_sorted]

        # Format datetime column names
        formatted = []
        for col in result_df.columns:
            if isinstance(col, pd.Timestamp):
                formatted.append(col.strftime("%Y-%m-%d"))
            else:
                formatted.append(col)
        result_df.columns = formatted

        return result_df

    # -------------------------------------------------------------------------
    # Excel formatting
    # -------------------------------------------------------------------------
    def _get_quarter_from_date(self, date_str):
        try:
            date_obj = pd.to_datetime(date_str) if isinstance(date_str, str) else date_str
            m = date_obj.month
            if m in [1, 2, 3]: return "Q1"
            if m in [4, 5, 6]: return "Q2"
            if m in [7, 8, 9]: return "Q3"
            if m in [10, 11, 12]: return "Q4"
            return ""
        except Exception:
            return ""

    def format_excel_sheet(self, writer, sheet_name, df):
        ws = writer.sheets[sheet_name]
        ws.insert_rows(1)

        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name != "Line_Item" and isinstance(col_name, str):
                quarter = self._get_quarter_from_date(col_name)
                cell = ws.cell(row=1, column=col_idx)
                cell.value = quarter

        q_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        q_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = q_fill
            cell.font = q_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        h_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        h_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[2]:
            cell.fill = h_fill
            cell.font = h_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column in ws.columns:
            max_len = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    txt = str(cell.value) if cell.value is not None else ""
                    if len(txt) > max_len:
                        max_len = len(txt)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        acct_fmt = "#,##0"
        for row in ws.iter_rows(min_row=3):
            for cell in row[1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = acct_fmt
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.freeze_panes = "B3"

    def export_to_excel(self, output_filename, start_year=2020):
        logger.info("=" * 60)
        logger.info(f"Starting comprehensive extraction for {self.company_name}")
        logger.info(f"Data range: {start_year} - Present")
        logger.info("=" * 60)

        df = self.extract_all_data(start_year)
        if df.empty:
            logger.warning("No data extracted!")
            return None

        df = self.calculate_q4_data(df)

        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
            # Raw data
            df.to_excel(writer, sheet_name="All Data - Raw", index=False)
            self.format_excel_sheet(writer, "All Data - Raw", df)

            # Income Statement
            logger.info("\nCreating Income Statement")
            income_pivot = self.create_statement_pivot(df, "income")
            if not income_pivot.empty:
                income_pivot.to_excel(writer, sheet_name="Income Statement - Quarterly", index=False)
                self.format_excel_sheet(writer, "Income Statement - Quarterly", income_pivot)

            # Balance Sheet
            logger.info("Creating Balance Sheet")
            balance_pivot = self.create_statement_pivot(df, "balance")
            if not balance_pivot.empty:
                balance_pivot.to_excel(writer, sheet_name="Balance Sheet - Quarterly", index=False)
                self.format_excel_sheet(writer, "Balance Sheet - Quarterly", balance_pivot)

            # Cash Flow
            logger.info("Creating Cash Flow Statement")
            cashflow_pivot = self.create_statement_pivot(df, "cashflow")
            if not cashflow_pivot.empty:
                cashflow_pivot.to_excel(writer, sheet_name="Cash Flow - Quarterly", index=False)
                self.format_excel_sheet(writer, "Cash Flow - Quarterly", cashflow_pivot)

            logger.info("\n" + "=" * 60)
            logger.info(f"✓ Export complete! File saved: {output_filename}")
            logger.info("=" * 60)

        return output_filename


def main():
    """Main execution"""
    YOUR_EMAIL = "brayden.joyce@doosan.com"  # clear identity per SEC guidance
    CIK = "0000018230"
    COMPANY_NAME = "Caterpillar Inc."
    TICKER = "cat"
    START_YEAR = 2020

    extractor = ComprehensiveXBRLExtractor(
        email=YOUR_EMAIL,
        cik=CIK,
        company_name=COMPANY_NAME,
        ticker=TICKER
    )

    output_file = extractor.export_to_excel(
        output_filename="caterpillar_financials.xlsx",
        start_year=START_YEAR
    )

    if output_file:
        print(f"\nSuccess! Complete financial data exported to: {output_file}")
        print(f"\nData range: {START_YEAR} - Present")
        print("\nThe Excel file contains:")
        print("  Income Statement - Quarterly")
        print("  Balance Sheet - Quarterly")
        print("  Cash Flow Statement - Quarterly")
        print("  All raw data")


if __name__ == "__main__":
    main()
