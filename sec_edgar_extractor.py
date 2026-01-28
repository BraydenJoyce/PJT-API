"""
SEC EDGAR Financial Data Extractor for Caterpillar Inc.
Extracts complete financial statement history and exports to Excel files.
"""

import requests
import pandas as pd
from datetime import datetime
import time
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class SECEdgarExtractor:
    """Extract financial data from SEC EDGAR API"""
    
    def __init__(self, email):
        """
        Initialize the extractor
        
        Args:
            email: Your email for SEC API user agent (required by SEC)
        """
        self.base_url = "https://data.sec.gov"
        self.headers = {
            'User-Agent': f'{email}',
            'Accept-Encoding': 'gzip, deflate',
            'Host': 'data.sec.gov'
        }
        self.cik = "0000018230"  # Caterpillar Inc. CIK
        self.company_name = "Caterpillar Inc."
        
    def get_company_facts(self):
        """
        Retrieve all company facts from SEC EDGAR API
        
        Returns:
            dict: Complete company facts data
        """
        url = f"{self.base_url}/api/xbrl/companyfacts/CIK{self.cik}.json"
        
        logger.info(f"Fetching company facts for {self.company_name} (CIK: {self.cik})")
        
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            time.sleep(0.1)  # Rate limiting
            
            data = response.json()
            logger.info("Successfully retrieved company facts")
            return data
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching data: {e}")
            raise
    
    def extract_financial_statement_data(self, facts_data, statement_type):
        """
        Extract specific financial statement data
        
        Args:
            facts_data: Complete company facts from API
            statement_type: Type of statement ('income', 'balance', 'equity')
            
        Returns:
            pd.DataFrame: Organized financial data
        """
        logger.info(f"Extracting {statement_type} statement data...")
        
        # Define line items for each statement type
        statement_items = self._get_statement_items(statement_type)
        
        # Extract US-GAAP facts
        us_gaap = facts_data.get('facts', {}).get('us-gaap', {})
        dei = facts_data.get('facts', {}).get('dei', {})
        
        all_records = []
        
        for item_name, item_label in statement_items.items():
            if item_name in us_gaap:
                item_data = us_gaap[item_name]
                units = item_data.get('units', {})
                
                # Process USD data
                if 'USD' in units:
                    for record in units['USD']:
                        all_records.append({
                            'Line_Item': item_label,
                            'XBRL_Tag': item_name,
                            'Value': record.get('val'),
                            'End_Date': record.get('end'),
                            'Start_Date': record.get('start'),
                            'Filed_Date': record.get('filed'),
                            'Form': record.get('form'),
                            'Fiscal_Year': record.get('fy'),
                            'Fiscal_Period': record.get('fp'),
                            'Accession_Number': record.get('accn'),
                            'Frame': record.get('frame')
                        })
        
        if not all_records:
            logger.warning(f"No data found for {statement_type} statement")
            return pd.DataFrame()
        
        df = pd.DataFrame(all_records)
        
        # Convert dates
        df['End_Date'] = pd.to_datetime(df['End_Date'])
        df['Filed_Date'] = pd.to_datetime(df['Filed_Date'])
        df['Start_Date'] = pd.to_datetime(df['Start_Date'])
        
        # Sort by end date and line item
        df = df.sort_values(['End_Date', 'Line_Item'], ascending=[False, True])
        
        logger.info(f"Extracted {len(df)} records for {statement_type} statement")
        return df
    
    def _get_statement_items(self, statement_type):
        """
        Get relevant line items for each statement type
        
        Args:
            statement_type: Type of financial statement
            
        Returns:
            dict: Mapping of XBRL tags to readable labels
        """
        if statement_type == 'income':
            return {
                #Sales and Revenues
                '': '',
                'Revenues': '     Sales of Machinery, Energy & Transportation',
                'Revenues': '     Revenues of Financial Products',                           #Same Tag?
                'Revenues': '     Total sales and revenues',

                #Operating Costs
                '': '',
                'CostOfRevenue': '     Cost of goods sold',
                'SellingGeneralAndAdministrativeExpense': '     SG&A Expenses',
                'ResearchAndDevelopmentExpense': '     R&D Expenses',
                'FinancingInterestExpense': '     Interest expense of Financial Products',             #Financial Products [Member]*****
                'OtherOperatingIncomeExpenseNet': '     Other operating (income) expenses',
                'CostsAndExpenses': '     Total operating costs',
                
                '': '',
                'OperatingIncomeLoss' : 'Operating Profit',

                '': '',
                'InterestExpenseNonoperating': '     Interest expense excluding Financial Products',
                'OtherNonoperatingIncomeExpense': '     Other income (expense)',

                '': '',
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': 'Consolidated profit before taxes',
                
                '': '',
                'IncomeTaxExpenseBenefit': '     Provision (benefit) for income taxes',
                'ProfitOfConsolidatedCompanies': '     Profit of consolidated companies',
            
                '': '',
                'IncomeLossFromEquityMethodInvestments': '     Equity in profit (loss) of unconsolidated affiliated companies',
                
                '': '',
                'ProfitLoss': 'Profit of consolidated and affiliated companies',

                '': '',
                'NetIncomeLossAttributableToNoncontrollingInterest': 'Less: Profit (loss) attributable to noncontrolling interests',

                '': '',
                'NetIncomeLossAvailableToCommonStockholdersBasic': 'Profit (Attributable to Common Stockholders)',
                
                # EPS
                '': '',
                'EarningsPerShareBasic': 'Profit per common share',

                '': '',
                'EarningsPerShareDiluted': 'Profit per common share - diluted',

                # Weighted average common shares outstanding (millions)
                '': '',
                'WeightedAverageNumberOfSharesOutstandingBasic': 'Shares Outstanding - Basic',
                'WeightedAverageNumberOfDilutedSharesOutstanding': 'Shares Outstanding - Diluted',
                    }
        
        elif statement_type == 'balance':
            return {
                #Assets
                #Current Assets
                'CashAndCashEquivalentsAtCarryingValue': '          Cash & Cash Equivalents',
                'AccountsReceivableNetCurrent': '          Receivables - trade and other',
                'NotesAndLoansReceivableNetCurrent': '          Receivables - finance',
                'PrepaidExpenseAndOtherAssetsCurrent': '          Prepaid Expenses And Other Assets Current',
                'InventoryNet': '          Inventories',
                'AssetsCurrent': '     Total Current Assets',

                'PropertyPlantAndEquipmentNet': '     Property, Plant, & Equipment - net',
                'AccountsReceivableNetNoncurrent': '     Long-term receivables - trade and other',
                'NotesAndLoansReceivableNetNoncurrent': '     Long-term receivables - finance',
                'NoncurrentDeferredAndRefundableIncomeTaxes': '     Noncurrent deferred and refundable income taxes',
                'IntangibleAssetsNetExcludingGoodwill': '     Intangible Assets',
                'Goodwill': '     Goodwill',
                'OtherAssetsNoncurrent': '     Other assets',
                'Assets': 'Total assets',

                #Liabilities              
                #Current liabilities:   
                #Short-term borrowings:
                'ShortTermBorrowings': '               Financial Products',
                'AccountsPayableCurrent': '          Accounts payable',
                'AccruedLiabilitiesCurrent': '          Accrued expenses',
                'EmployeeRelatedLiabilitiesCurrent': '          Accrued wages: salaries and employee benefits',
                'ContractWithCustomerLiabilityCurrent': '          Customer advances',
                'DividendsPayableCurrent': '          Dividends payable',
                'OtherLiabilitiesCurrent': '          Other current liabilities',

                #Long-term debt due within one year:
                'LongTermDebtAndCapitalLeaseObligationsCurrent': '               Machinery: Energy & Transportation', #Same Tag?
                'LongTermDebtAndCapitalLeaseObligationsCurrent': '               Financial Products',
                'LiabilitiesCurrent': '     Total current liabilities',

                #Long-term debt due after one year:
                'LongTermDebtAndCapitalLeaseObligations': '               Machinery: Energy & Transportation', #Same Tag?
                'LongTermDebtAndCapitalLeaseObligations': '               Financial Products',
                'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent': '     Liability for postemployment benefits',
                'OtherLiabilitiesNoncurrent': '     Other liabilities',
                'Total Liabilities': 'Liabilities',

                #Shareholders' Equity
                'CommonStocksIncludingAdditionalPaidInCapital': '     Issued shares at paid-in amount',
                'TreasuryStockValue': '     Treasury stock at cost',
                'RetainedEarningsAccumulatedDeficit': '     Profit employed in the business',
                'AccumulatedOtherComprehensiveIncomeLossNetOfTax': '     Accumulated other comprehensive income (loss)',
                'MinorityInterest': '     Noncontrolling interests',
                'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest': 'Total shareholders’ equity',
                'Total liabilities and shareholders’ equity': 'LiabilitiesAndStockholdersEquity',
                }
        
        elif statement_type == 'cashflow':
            return {
                # Operating Activities
                'ProfitLoss': '     Profit of consolidated and affiliated companies',
                #Adjustments to reconcile profit to net cash provided by operating activities
                'DepreciationDepletionAndAmortization': '          Depreciation and Amortization',
                'DeferredIncomeTaxExpenseBenefit': '          Provision (benefit) for deferred income taxes',
                'NonCashGainLossOnDivestiture': '          (Gain) loss on divestiture',
                'OtherNoncashIncomeExpense': '          Other',
                #Changes in assets and liabilities, net of acquisitions and divestitures
                'IncreaseDecreaseInReceivables': '          Receivables – trade and other',
                'Inventories': 'IncreaseDecreaseInInventories',
                'IncreaseDecreaseInAccountsPayable': '          Accounts payable',
                'IncreaseDecreaseInAccruedLiabilities': '           Accrued expenses',
                'IncreaseDecreaseInEmployeeRelatedLiabilities': '          Accrued wages, salaries and employee benefits',
                'IncreaseDecreaseInContractWithCustomerLiability': '          Customer advances',
                'IncreaseDecreaseInOtherOperatingAssets': '          Other assets - net',
                'IncreaseDecreaseInOtherOperatingLiabilities': '          Other liabilities - net',
                'NetCashProvidedByUsedInOperatingActivities': 'Net cash provided by (used for) operating activities',
                
                # Investing Activities
                'PaymentsToAcquirePropertyPlantAndEquipment': '     Capital expenditures – excluding equipment leased to others',
                'PaymentsToAcquireEquipmentOnLease': '     Expenditures for equipment leased to others',
                'ProceedsFromSaleOfPropertyPlantAndEquipment': '     Proceeds from disposals of leased assets and property, plant and equipment',
                'PaymentsToAcquireFinanceReceivables': '     Additions to finance receivables',
                'ProceedsFromCollectionOfFinanceReceivables': '     Collections of finance receivables',
                'ProceedsFromSaleOfFinanceReceivables': '     Proceeds from sale of finance receivables',
                'PaymentsToAcquireBusinessesNetOfCashAcquired': '     Investments and acquisitions (net of cash acquired)',
                'ProceedsFromDivestitureOfBusinessesNetOfCashDivested': '     Proceeds from sale of businesses and investments (net of cash sold)',
                'ProceedsFromSaleAndMaturityOfMarketableSecurities': '      Proceeds from maturities and sale of securities',
                'PaymentsToAcquireMarketableSecurities': '      Investments in securities',
                'PaymentsForProceedsFromOtherInvestingActivities': '     Other – net',
                'NetCashProvidedByUsedInInvestingActivities': 'Net cash provided by (used for) investing activities',
                
                # Cash flow from financing activities
                'PaymentsOfDividendsCommonStock': '     Dividends paid',
                'ProceedsFromIssuanceOrSaleOfEquity': '     Common stock issued, and other stock compensation transactions, net',
                'PaymentsForRepurchaseOfCommonStock': '     Payments to purchase common stock',
                'PaymentsForExciseTaxOnPurchaseOfCommonStock': '     Excise tax paid on purchases of common stock',
                #Proceeds from debt issued (original maturities greater than three months)
                'ProceedsFromDebtMaturingInMoreThanThreeMonths': '          - Machinery, Energy & Transportation',
                'ProceedsFromDebtMaturingInMoreThanThreeMonths': '          - Financial Products',
                #Payments on debt (original maturities greater than three months)
                'RepaymentsOfDebtMaturingInMoreThanThreeMonths': '          - Machinery, Energy & Transportation',
                'RepaymentsOfDebtMaturingInMoreThanThreeMonths': '          - Financial Products',
                'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess': '     Short-term borrowings – net (original maturities three months or less)',
                'NetCashProvidedByUsedInFinancingActivities': 'Net cash provided by (used for) financing activities',
                'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': 'Effect of exchange rate changes on cash',
                #
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect': 'Increase (decrease) in cash, cash equivalents and restricted cash',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': 'Cash, cash equivalents and restricted cash at beginning of period',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': 'Cash, cash equivalents and restricted cash at end of period',
            }
        
        else:
            logger.warning(f"Unknown statement type: {statement_type}")
            return {}
    
    def create_pivot_table(self, df, statement_type):
        """
        Create a pivot table view of the financial data
        
        Args:
            df: DataFrame with financial data
            statement_type: Type of statement
            
        Returns:
            pd.DataFrame: Pivoted data showing values across periods
        """
        if df.empty:
            return pd.DataFrame()
        
        # Filter to annual and Quarterly reports only (10-K and 10-Q)
        annual_df = df[df['Form'].isin(['10-K', '10-Q'])].copy()
        
        if annual_df.empty:
            logger.warning(f"No 10-K or 10-Q data found for {statement_type}")
            return df
        
        # For income statement, filter to only 3-month periods (approximately 90-100 days)
        if statement_type == 'income':
            annual_df['Period_Length'] = (annual_df['End_Date'] - annual_df['Start_Date']).dt.days
            annual_df = annual_df[annual_df['Period_Length'] <= 100]
            if annual_df.empty:
                logger.warning(f"No 3-month period data found for {statement_type}")
                return pd.DataFrame()
        
        # Create pivot table
        pivot = annual_df.pivot_table(
            index='Line_Item',
            columns='End_Date',
            values='Value',
            aggfunc='first'
        )
        
        # Sort columns by date (most recent first)
        pivot = pivot[sorted(pivot.columns, reverse=True)]
        
        # Format column names
        pivot.columns = [col.strftime('%Y-%m-%d') if isinstance(col, datetime) else str(col)
                        for col in pivot.columns]
        
        # Preserve the line-item order as declared in _get_statement_items
        desired_order = list(self._get_statement_items(statement_type).values())
        ordered_index = [lbl for lbl in desired_order if lbl in pivot.index]
        ordered_index += [lbl for lbl in pivot.index if lbl not in ordered_index]
        pivot = pivot.reindex(ordered_index)
        
        return pivot
    
    def format_excel_sheet(self, writer, sheet_name, df):
        """
        Apply formatting to Excel sheet
        
        Args:
            writer: ExcelWriter object
            sheet_name: Name of the sheet
            df: DataFrame that was written
        """
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header formatting
        header_format = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        # Apply header formatting to the actual first row in the worksheet
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.fill = header_format
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    # measure string length of the cell value safely
                    text = str(cell.value) if cell.value is not None else ""
                    if len(text) > max_length:
                        max_length = len(text)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply accounting number format to numeric cells (exclude first column which holds line items/index)
        accounting_format = '#,##0'
        for row in worksheet.iter_rows(min_row=2):  # skip header row
            for cell in row[1:]:  # skip first column (A)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = accounting_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Freeze first column (A) and header row simultaneously
        worksheet.freeze_panes = 'B2'
    
    def export_to_excel(self, output_filename='caterpillar_financials.xlsx'):
        """
        Main function to extract all data and export to Excel
        
        Args:
            output_filename: Name of output Excel file
        """
        logger.info("="*60)
        logger.info("Starting SEC EDGAR data extraction for Caterpillar Inc.")
        logger.info("="*60)
        
        # Get all company facts
        facts_data = self.get_company_facts()
        
        # Create Excel writer
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            
            # 1. Income Statement
            logger.info("\n" + "="*60)
            logger.info("PROCESSING INCOME STATEMENT")
            logger.info("="*60)
            income_df = self.extract_financial_statement_data(facts_data, 'income')
            if not income_df.empty:
                # Raw data
                income_df.to_excel(writer, sheet_name='Income Statement - Raw', index=False)
                self.format_excel_sheet(writer, 'Income Statement - Raw', income_df)
                
                # Pivot view
                income_pivot = self.create_pivot_table(income_df, 'income')
                if not income_pivot.empty:
                    income_pivot.to_excel(writer, sheet_name='Income Statement - Quarterly')
                    self.format_excel_sheet(writer, 'Income Statement - Quarterly', income_pivot)
            
            # 2. Balance Sheet
            logger.info("\n" + "="*60)
            logger.info("PROCESSING BALANCE SHEET")
            logger.info("="*60)
            balance_df = self.extract_financial_statement_data(facts_data, 'balance')
            if not balance_df.empty:
                # Raw data
                balance_df.to_excel(writer, sheet_name='Balance Sheet - Raw', index=False)
                self.format_excel_sheet(writer, 'Balance Sheet - Raw', balance_df)
                
                # Pivot view
                balance_pivot = self.create_pivot_table(balance_df, 'balance')
                if not balance_pivot.empty:
                    balance_pivot.to_excel(writer, sheet_name='Balance Sheet - Quarterly')
                    self.format_excel_sheet(writer, 'Balance Sheet - Quarterly', balance_pivot)
            
            # 3. Cash Flow Statement
            logger.info("\n" + "="*60)
            logger.info("PROCESSING CASH FLOW STATEMENT")
            logger.info("="*60)
            cashflow_df = self.extract_financial_statement_data(facts_data, 'cashflow')
            if not cashflow_df.empty:
                # Raw data
                cashflow_df.to_excel(writer, sheet_name='Cash Flow - Raw', index=False)
                self.format_excel_sheet(writer, 'Cash Flow - Raw', cashflow_df)
                
                # Pivot view
                cashflow_pivot = self.create_pivot_table(cashflow_df, 'cashflow')
                if not cashflow_pivot.empty:
                    cashflow_pivot.to_excel(writer, sheet_name='Cash Flow - Quarterly')
                    self.format_excel_sheet(writer, 'Cash Flow - Quarterly', cashflow_pivot)
        
        logger.info("\n" + "="*60)
        logger.info(f"✓ Export complete! File saved: {output_filename}")
        logger.info("="*60)
        
        return output_filename


def main():
    """Main execution function"""
    
    # IMPORTANT: Replace with your email address
    YOUR_EMAIL = "brayden.joyce@doosan.com"
    
    if YOUR_EMAIL == "your.email@example.com":
        print("\n" + "!"*60)
        print("IMPORTANT: Please update YOUR_EMAIL in the script (line 451)")
        print("The SEC requires a valid email in the User-Agent header")
        print("!"*60 + "\n")
        return
    
    # Create extractor instance
    extractor = SECEdgarExtractor(email=YOUR_EMAIL)
    
    # Extract and export data
    output_file = extractor.export_to_excel('caterpillar_financials.xlsx')
    
    print(f"\n✓ Success! Financial data exported to: {output_file}")
    print("\nThe Excel file contains:")
    print("  • Income Statement (Raw data + Quarterly pivot)")
    print("  • Balance Sheet (Raw data + Quarterly pivot)")
    print("  • Cash Flow Statement (Raw data + Quarterly pivot)")
    print("\nEach statement has complete historical data from SEC EDGAR!")


if __name__ == "__main__":
    main()
