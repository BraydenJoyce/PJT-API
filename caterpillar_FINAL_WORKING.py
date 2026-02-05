#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SEC EDGAR XBRL Parser - Complete Financial Statement Extractor
Extracts both consolidated and segment-level data with Q4 calculations

This version ensures:
- Q4 calculations are done **only for Consolidated** (never for segments).
- Segment tabs exclude any calculated rows; they show only reported quarterlies.
- Discrete-quarter conversion for YTD 10-Qs is limited to Consolidated only.
"""

import requests
import pandas as pd
from datetime import datetime
import time
import logging
import xml.etree.ElementTree as ET
from openpyxl import load_workbook  # noqa: F401 (kept for parity; not used directly)
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import OrderedDict

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ComprehensiveXBRLExtractor:
    """Extract complete financial statements with segment breakdowns"""

    def __init__(self, email, cik, company_name, ticker):
        """
        Initialize the extractor
        Args:
            email: Your email for SEC API user agent
            cik: Company CIK number (with leading zeros)
            company_name: Company name for logging
            ticker: Company ticker symbol (lowercase, for URL construction)
        """
        self.base_url = "https://data.sec.gov"
        self.sec_archives = "https://www.sec.gov/Archives/edgar/data"
        self.headers = {
            'User-Agent': f'{email}',
            'Accept-Encoding': 'gzip, deflate'
        }
        self.cik = cik
        self.cik_int = str(int(cik))
        self.company_name = company_name
        self.ticker = ticker.lower()

        # XBRL namespaces
        self.namespaces = {
            'xbrli': 'http://www.xbrl.org/2003/instance',
            'xbrldi': 'http://xbrl.org/2006/xbrldi',
        }

    def _get_statement_items(self, statement_type):
        """Get line items in proper order for each statement"""
        if statement_type == 'income':
            return OrderedDict([
                # Sales and Revenues
                ('Revenues_MET', ' Sales of Machinery, Energy & Transportation'),
                ('Revenues_FP', ' Revenues of Financial Products'),
                ('Revenues', ' Total sales and revenues'),
                # Operating Costs
                ('CostOfRevenue', ' Cost of goods sold'),
                ('SellingGeneralAndAdministrativeExpense', ' SG&A Expenses'),
                ('ResearchAndDevelopmentExpense', ' R&D Expenses'),
                ('FinancingInterestExpense_FP', ' Interest expense of Financial Products'),
                ('OtherOperatingIncomeExpenseNet', ' Other operating (income) expenses'),
                ('CostsAndExpenses', ' Total operating costs'),
                ('OperatingIncomeLoss', 'Operating Profit'),
                ('InterestExpenseNonoperating_EXFP', ' Interest expense excluding Financial Products'),
                ('OtherNonoperatingIncomeExpense', ' Other income (expense)'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'Consolidated profit before taxes'),
                ('IncomeTaxExpenseBenefit', ' Provision (benefit) for income taxes'),
                ('ProfitOfConsolidatedCompanies', ' Profit of consolidated companies'),
                ('IncomeLossFromEquityMethodInvestments', ' Equity in profit (loss) of unconsolidated affiliated companies'),
                ('ProfitLoss', ' Profit of consolidated and affiliated companies'),
                ('NetIncomeLossAttributableToNoncontrollingInterest', ' Less: Profit (loss) attributable to noncontrolling interests'),
                ('NetIncomeLossAvailableToCommonStockholdersBasic', ' Profit (Attributable to Common Stockholders)'),
                # EPS
                ('EarningsPerShareBasic', 'Profit per common share'),
                ('EarningsPerShareDiluted', 'Profit per common share - diluted'),
                # Shares Outstanding
                ('WeightedAverageNumberOfSharesOutstandingBasic', 'Shares Outstanding - Basic'),
                ('WeightedAverageNumberOfDilutedSharesOutstanding', 'Shares Outstanding - Diluted'),
            ])
        elif statement_type == 'balance':
            return OrderedDict([
                # Assets
                # Current Assets
                ('CashAndCashEquivalentsAtCarryingValue', ' Cash & Cash Equivalents'),
                ('AccountsReceivableNetCurrent', ' Receivables - trade and other'),
                ('NotesAndLoansReceivableNetCurrent', ' Receivables - finance'),
                ('PrepaidExpenseAndOtherAssetsCurrent', ' Prepaid Expenses And Other Assets Current'),
                ('InventoryNet', ' Inventories'),
                ('AssetsCurrent', ' Total Current Assets'),
                ('PropertyPlantAndEquipmentNet', ' Property, Plant, & Equipment - net'),
                ('AccountsReceivableNetNoncurrent', ' Long-term receivables - trade and other'),
                ('NotesAndLoansReceivableNetNoncurrent', ' Long-term receivables - finance'),
                ('NoncurrentDeferredAndRefundableIncomeTaxes', ' Noncurrent deferred and refundable income taxes'),
                ('IntangibleAssetsNetExcludingGoodwill', ' Intangible Assets'),
                ('Goodwill', ' Goodwill'),
                ('OtherAssetsNoncurrent', ' Other assets'),
                ('Assets', 'Total assets'),
                # Liabilities
                # Current liabilities
                ('ShortTermBorrowings_FinancialProducts', ' Financial Products'),
                ('AccountsPayableCurrent', ' Accounts payable'),
                ('AccruedLiabilitiesCurrent', ' Accrued expenses'),
                ('EmployeeRelatedLiabilitiesCurrent', ' Accrued wages: salaries and employee benefits'),
                ('ContractWithCustomerLiabilityCurrent', ' Customer advances'),
                ('DividendsPayableCurrent', ' Dividends payable'),
                ('OtherLiabilitiesCurrent', ' Other current liabilities'),
                # Long-term debt
                ('LongTermDebtAndCapitalLeaseObligations_MET', ' Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligations_FP', ' Financial Products'),
                ('LiabilitiesCurrent', ' Total current liabilities'),
                ('LongTermDebtAndCapitalLeaseObligations_MET', ' Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligations_FP', ' Financial Products'),
                ('PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent', ' Liability for postemployment benefits'),
                ('OtherLiabilitiesNoncurrent', ' Other liabilities'),
                ('Liabilities', 'Total Liabilities'),
                # Shareholders' Equity
                ('CommonStocksIncludingAdditionalPaidInCapital', ' Issued shares at paid-in amount'),
                ('TreasuryStockValue', ' Treasury stock at cost'),
                ('RetainedEarningsAccumulatedDeficit', ' Profit employed in the business'),
                ('AccumulatedOtherComprehensiveIncomeLossNetOfTax', ' Accumulated other comprehensive income (loss)'),
                ('MinorityInterest', ' Noncontrolling interests'),
                ('StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', "Total shareholders' equity"),
                ('LiabilitiesAndStockholdersEquity', "Total liabilities and shareholders' equity"),
            ])
        elif statement_type == 'cashflow':
            return OrderedDict([
                # Operating Activities
                ('ProfitLoss', ' Profit of consolidated and affiliated companies'),
                # Adjustments
                ('DepreciationDepletionAndAmortization', ' Depreciation and Amortization'),
                ('DeferredIncomeTaxExpenseBenefit', ' Provision (benefit) for deferred income taxes'),
                ('NonCashGainLossOnDivestiture', ' (Gain) loss on divestiture'),
                ('OtherNoncashIncomeExpense', ' Other'),
                # Changes in assets and liabilities
                ('IncreaseDecreaseInReceivables', ' Receivables – trade and other'),
                ('IncreaseDecreaseInInventories', ' Inventories'),
                ('IncreaseDecreaseInAccountsPayable', ' Accounts payable'),
                ('IncreaseDecreaseInAccruedLiabilities', ' Accrued expenses'),
                ('IncreaseDecreaseInEmployeeRelatedLiabilities', ' Accrued wages, salaries and employee benefits'),
                ('IncreaseDecreaseInContractWithCustomerLiability', ' Customer advances'),
                ('IncreaseDecreaseInOtherOperatingAssets', ' Other assets - net'),
                ('IncreaseDecreaseInOtherOperatingLiabilities', ' Other liabilities - net'),
                ('NetCashProvidedByUsedInOperatingActivities', ' Net cash provided by (used for) operating activities'),
                # Investing Activities
                ('PaymentsToAcquirePropertyPlantAndEquipment', ' Capital expenditures – excluding equipment leased to others'),
                ('PaymentsToAcquireEquipmentOnLease', ' Expenditures for equipment leased to others'),
                ('ProceedsFromSaleOfPropertyPlantAndEquipment', ' Proceeds from disposals of leased assets and property, plant and equipment'),
                ('PaymentsToAcquireFinanceReceivables', ' Additions to finance receivables'),
                ('ProceedsFromCollectionOfFinanceReceivables', ' Collections of finance receivables'),
                ('ProceedsFromSaleOfFinanceReceivables', ' Proceeds from sale of finance receivables'),
                ('PaymentsToAcquireBusinessesNetOfCashAcquired', ' Investments and acquisitions (net of cash acquired)'),
                ('ProceedsFromDivestitureOfBusinessesNetOfCashDivested', ' Proceeds from sale of businesses and investments (net of cash sold)'),
                ('ProceedsFromSaleAndMaturityOfMarketableSecurities', ' Proceeds from maturities and sale of securities'),
                ('PaymentsToAcquireMarketableSecurities', ' Investments in securities'),
                ('PaymentsForProceedsFromOtherInvestingActivities', ' Other – net'),
                ('NetCashProvidedByUsedInInvestingActivities', ' Net cash provided by (used for) investing activities'),
                # Financing Activities
                ('PaymentsOfDividendsCommonStock', ' Dividends paid'),
                ('ProceedsFromIssuanceOrSaleOfEquity', ' Common stock issued, and other stock compensation transactions, net'),
                ('PaymentsForRepurchaseOfCommonStock', ' Payments to purchase common stock'),
                ('PaymentsForExciseTaxOnPurchaseOfCommonStock', ' Excise tax paid on purchases of common stock'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_MET', ' - Machinery, Energy & Transportation'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_FP', ' - Financial Products'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_MET', 'Payments on debt (Machinery, Energy & Transportation)'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_FP', 'Payments on debt (Financial Products)'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths', ' Proceeds from debt issued (Financial Products)'),
                ('ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess', ' Short-term borrowings – net (original maturities three months or less)'),
                ('NetCashProvidedByUsedInFinancingActivities', ' Net cash provided by (used for) financing activities'),
                ('EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents', 'Effect of exchange rate changes on cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect', ' Increase (decrease) in cash, cash equivalents and restricted cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning', 'Cash, cash equivalents and restricted cash at beginning of period'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End', 'Cash, cash equivalents and restricted cash at end of period'),
            ])
        elif statement_type == 'operatingsegments':
            return OrderedDict([
                ('Revenues_OpSeg', 'Sales and Revenues'),
                ('CostOfRevenue_OpSeg', 'Cost of Goods Sold'),
                ('SellingGeneralAndAdministrativeResearchAndDevelopment_OpSeg', 'SG&A/R&D'),
                ('SegmentReportingOtherItemAmount_OpSeg', 'Other segment items'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments_OpSeg', 'Profit'),
                ('DepreciationDepletionAndAmortization_OpSeg', 'Depreciation and Amortization'),
                ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets_OpSeg', 'Capital Expenditures'),
            ])
        elif statement_type == 'constructionindustries':
            return OrderedDict([
                ('Revenues_CI', 'Sales and Revenues'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments_CI', 'Profit'),
                ('DepreciationDepletionAndAmortization_CI', 'Depreciation and Amortization'),
                ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets_CI', 'Capital Expenditures'),
                # Note: COGS, SG&A not reported at individual segment level - see Operating Segments sheet
            ])
        elif statement_type == 'resourceindustries':
            return OrderedDict([
                ('Revenues_RI', 'Sales and Revenues'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments_RI', 'Profit'),
                ('DepreciationDepletionAndAmortization_RI', 'Depreciation and Amortization'),
                ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets_RI', 'Capital Expenditures'),
            ])
        elif statement_type == 'energyandtransportation':
            return OrderedDict([
                ('Revenues_ET', 'Sales and Revenues'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments_ET', 'Profit'),
                ('DepreciationDepletionAndAmortization_ET', 'Depreciation and Amortization'),
                ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets_ET', 'Capital Expenditures'),
            ])
        elif statement_type == 'financialproducts':
            return OrderedDict([
                ('Revenues_FPS', 'Sales and Revenues'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments_FPS', 'Profit'),
                ('DepreciationDepletionAndAmortization_FPS', 'Depreciation and Amortization'),
                ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets_FPS', 'Capital Expenditures'),
            ])
        else:
            logger.warning(f"Unknown statement type: {statement_type}")
            return OrderedDict()

    # --- ENHANCED: More robust tag matching for income statement ---
    def _get_income_tag_candidates(self):
        """
        Return mapping from our income statement line keys to lists of acceptable XBRL tags.
        This helps when companies use different but equivalent tags.
        """
        return {
            'Revenues': ['Revenues','SalesRevenueNet','SalesAndRevenue','SalesRevenueGoodsNet','SalesRevenueServicesNet','RevenueFromContractWithCustomerExcludingAssessedTax'],
            'SellingGeneralAndAdministrativeExpense': ['SellingGeneralAndAdministrativeExpense'],
            'ResearchAndDevelopmentExpense': ['ResearchAndDevelopmentExpense'],
            'OtherOperatingIncomeExpenseNet': ['OtherOperatingIncomeExpenseNet','OtherOperatingIncomeExpense'],
            'CostsAndExpenses': ['CostsAndExpenses','OperatingExpenses'],
            'OperatingIncomeLoss': ['OperatingIncomeLoss'],
            'InterestExpenseNonoperating': ['InterestExpenseNonoperating','InterestAndDebtExpense'],
            'InterestExpenseNonoperating_EXFP': ['InterestExpense_EXFP','InterestExpenseExcludingFinancialProducts'],
            'OtherNonoperatingIncomeExpense': ['OtherNonoperatingIncomeExpense','NonoperatingIncomeExpense'],
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': ['IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments','IncomeLossFromContinuingOperationsBeforeIncomeTaxes'],
            'IncomeTaxExpenseBenefit': ['IncomeTaxExpenseBenefit'],
            'IncomeLossFromEquityMethodInvestments': ['IncomeLossFromEquityMethodInvestments'],
            'ProfitLoss': ['ProfitLoss','NetIncomeLoss'],
            'NetIncomeLossAttributableToNoncontrollingInterest': ['NetIncomeLossAttributableToNoncontrollingInterest'],
            'NetIncomeLossAvailableToCommonStockholdersBasic': ['NetIncomeLossAvailableToCommonStockholdersBasic'],
            'EarningsPerShareBasic': ['EarningsPerShareBasic'],
            'EarningsPerShareDiluted': ['EarningsPerShareDiluted'],
            'WeightedAverageNumberOfSharesOutstandingBasic': ['WeightedAverageNumberOfSharesOutstandingBasic'],
            'WeightedAverageNumberOfDilutedSharesOutstanding': ['WeightedAverageNumberOfDilutedSharesOutstanding'],
            'FinancingInterestExpense_FP': ['InterestExpense','InterestAndDebtExpense','InterestExpenseOfFinancialProducts'],
            'CashAndCashEquivalentsAtCarryingValue' : ['CashCashEquivalentsAndShortTermInvestments']
        }

    def _get_balance_tag_candidates(self):
        """
        Alias map for balance sheet line items (common US-GAAP variants).
        Used to locate equivalent concepts when filers use alternative tags.
        """
        return {
            'CashAndCashEquivalentsAtCarryingValue': ['CashAndCashEquivalentsAtCarryingValue','CashCashEquivalentsAtCarryingValue','CashAndCashEquivalentsFairValueDisclosure','CashCashEquivalentsAndShortTermInvestments'],
            'AccountsReceivableNetCurrent': ['AccountsReceivableNetCurrent','AccountsReceivableNet','TradeAccountsReceivableNet','ReceivablesNetCurrent'],
            'NotesAndLoansReceivableNetCurrent': ['NotesAndLoansReceivableNetCurrent','LoansAndLeasesReceivableNetCurrent','FinanceReceivableCurrent'],
            'PrepaidExpenseAndOtherAssetsCurrent': ['PrepaidExpenseAndOtherAssetsCurrent','PrepaidExpenseCurrent','OtherAssetsCurrent'],
            'InventoryNet': ['InventoryNet','InventoriesNet','InventoryFinishedGoods'],
            'AssetsCurrent': ['AssetsCurrent','CurrentAssets'],
            'PropertyPlantAndEquipmentNet': ['PropertyPlantAndEquipmentNet','PropertyPlantAndEquipmentAndFinanceLeaseRightOfUseAssetAfterAccumulatedDepreciationAndAmortization'],
            'AccountsReceivableNetNoncurrent': ['AccountsReceivableNetNoncurrent','AccountsAndLoansReceivableNetNoncurrent'],
            'NotesAndLoansReceivableNetNoncurrent': ['NotesAndLoansReceivableNetNoncurrent','LoansAndLeasesReceivableNetNoncurrent','FinanceReceivableNoncurrent'],
            'NoncurrentDeferredAndRefundableIncomeTaxes': ['NoncurrentDeferredAndRefundableIncomeTaxes','DeferredTaxAssetsNetNoncurrent'],
            'IntangibleAssetsNetExcludingGoodwill': ['IntangibleAssetsNetExcludingGoodwill','FiniteLivedIntangibleAssetsNet'],
            'Goodwill': ['Goodwill'],
            'OtherAssetsNoncurrent': ['OtherAssetsNoncurrent','OtherNoncurrentAssets'],
            'Assets': ['Assets','TotalAssets'],
            'ShortTermBorrowings_FinancialProducts': ['ShortTermBorrowings','CommercialPaper','ShortTermDebt'],
            'AccountsPayableCurrent': ['AccountsPayableCurrent','TradeAccountsPayableCurrent'],
            'AccruedLiabilitiesCurrent': ['AccruedLiabilitiesCurrent','AccruedExpensesCurrent'],
            'EmployeeRelatedLiabilitiesCurrent': ['EmployeeRelatedLiabilitiesCurrent','AccruedCompensationCurrent'],
            'ContractWithCustomerLiabilityCurrent': ['ContractWithCustomerLiabilityCurrent','DeferredRevenueCurrent','CustomerAdvancesCurrent'],
            'DividendsPayableCurrent': ['DividendsPayableCurrent'],
            'OtherLiabilitiesCurrent': ['OtherLiabilitiesCurrent','OtherCurrentLiabilities'],
            'LiabilitiesCurrent': ['LiabilitiesCurrent','CurrentLiabilities'],
            'LongTermDebtAndCapitalLeaseObligations': ['LongTermDebtAndCapitalLeaseObligations','LongTermDebtNoncurrent','LongTermDebt'],
            'LongTermDebtAndCapitalLeaseObligationsCurrent': ['LongTermDebtAndCapitalLeaseObligationsCurrent','LongTermDebtCurrent'],
            'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent': ['PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent','EmployeeRelatedLiabilitiesNoncurrent'],
            'OtherLiabilitiesNoncurrent': ['OtherLiabilitiesNoncurrent','OtherNoncurrentLiabilities'],
            'Liabilities': ['Liabilities','TotalLiabilities'],
            'CommonStocksIncludingAdditionalPaidInCapital': ['CommonStocksIncludingAdditionalPaidInCapital','CommonStockValue','AdditionalPaidInCapital'],
            'TreasuryStockValue': ['TreasuryStockValue','TreasuryStockCommon'],
            'RetainedEarningsAccumulatedDeficit': ['RetainedEarningsAccumulatedDeficit','RetainedEarnings'],
            'AccumulatedOtherComprehensiveIncomeLossNetOfTax': ['AccumulatedOtherComprehensiveIncomeLossNetOfTax','AccumulatedOtherComprehensiveIncome'],
            'MinorityInterest': ['MinorityInterest','NoncontrollingInterest'],
            'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest': ['StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest','StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest','StockholdersEquity'],
            'LiabilitiesAndStockholdersEquity': ['LiabilitiesAndStockholdersEquity','LiabilitiesAndStockholdersEquity']
        }

    def _get_cashflow_tag_candidates(self):
        """
        Alias map for cash flow line items (common US-GAAP variants + near-equivalents).
        Expand with company extensions (e.g., cat:*) after inspecting df['tag'].
        """
        return {
            'ProfitLoss': ['ProfitLoss', 'NetIncomeLoss'],
            'DepreciationDepletionAndAmortization': ['DepreciationDepletionAndAmortization','DepreciationAndAmortization','DepreciationAmortizationAndAccretionNet'],
            'DeferredIncomeTaxExpenseBenefit': ['DeferredIncomeTaxExpenseBenefit','DeferredIncomeTaxesAndTaxCredits'],
            'NonCashGainLossOnDivestiture': ['GainLossOnSaleOfBusiness','NoncashOrPartNoncashAcquisitionOrDisposition'],
            'OtherNoncashIncomeExpense': ['OtherNoncashIncomeExpense','OtherNoncashOperatingActivities'],
            'IncreaseDecreaseInReceivables': ['IncreaseDecreaseInReceivables','IncreaseDecreaseInAccountsReceivable','IncreaseDecreaseInAccountsAndNotesReceivable','IncreaseDecreaseInTradeAccountsReceivable'],
            'IncreaseDecreaseInInventories': ['IncreaseDecreaseInInventories','IncreaseDecreaseInInventory'],
            'IncreaseDecreaseInAccountsPayable': ['IncreaseDecreaseInAccountsPayable','IncreaseDecreaseInTradeAccountsPayable'],
            'IncreaseDecreaseInAccruedLiabilities': ['IncreaseDecreaseInAccruedLiabilities','IncreaseDecreaseInAccruedExpensesAndOtherLiabilities'],
            'IncreaseDecreaseInEmployeeRelatedLiabilities': ['IncreaseDecreaseInEmployeeRelatedLiabilities','IncreaseDecreaseInAccruedPayrollAndBenefits'],
            'IncreaseDecreaseInContractWithCustomerLiability': ['IncreaseDecreaseInContractWithCustomerLiability','IncreaseDecreaseInDeferredRevenue','IncreaseDecreaseInCustomerAdvances'],
            'IncreaseDecreaseInOtherOperatingAssets': ['IncreaseDecreaseInOtherOperatingAssets','IncreaseDecreaseInOtherAssets'],
            'IncreaseDecreaseInOtherOperatingLiabilities': ['IncreaseDecreaseInOtherOperatingLiabilities','IncreaseDecreaseInOtherLiabilities'],
            'NetCashProvidedByUsedInOperatingActivities': ['NetCashProvidedByUsedInOperatingActivities','NetCashProvidedByUsedInOperatingActivitiesContinuingOperations'],
            'PaymentsToAcquirePropertyPlantAndEquipment': ['PaymentsToAcquirePropertyPlantAndEquipment','CapitalExpenditures'],
            'PaymentsToAcquireEquipmentOnLease': ['PaymentsToAcquireEquipmentOnLease','PaymentsToAcquireEquipmentLeasedToOthers'],
            'ProceedsFromSaleOfPropertyPlantAndEquipment': ['ProceedsFromSaleOfPropertyPlantAndEquipment','ProceedsFromSalesOfPropertyPlantAndEquipment','ProceedsFromDisposalsOfLeasedAssetsAndPropertyPlantAndEquipment'],
            'PaymentsToAcquireFinanceReceivables': ['PaymentsToAcquireFinanceReceivables','IncreaseInFinanceReceivables'],
            'ProceedsFromCollectionOfFinanceReceivables': ['ProceedsFromCollectionOfFinanceReceivables','CollectionsOfFinanceReceivables'],
            'ProceedsFromSaleOfFinanceReceivables': ['ProceedsFromSaleOfFinanceReceivables'],
            'PaymentsToAcquireBusinessesNetOfCashAcquired': ['PaymentsToAcquireBusinessesNetOfCashAcquired','PaymentsForBusinessCombinationsNetOfCashAcquired'],
            'ProceedsFromDivestitureOfBusinessesNetOfCashDivested': ['ProceedsFromDivestitureOfBusinessesNetOfCashDivested','ProceedsFromSaleOfBusinessNetOfCashDisposed'],
            'ProceedsFromSaleAndMaturityOfMarketableSecurities': ['ProceedsFromSaleAndMaturityOfMarketableSecurities','ProceedsFromMaturiesSaleOfInvestments','ProceedsFromMaturitiesSaleOfInvestments'],
            'PaymentsToAcquireMarketableSecurities': ['PaymentsToAcquireMarketableSecurities','PaymentsToAcquireInvestments'],
            'PaymentsForProceedsFromOtherInvestingActivities': ['PaymentsForProceedsFromOtherInvestingActivities','OtherInvestingActivitiesNet'],
            'NetCashProvidedByUsedInInvestingActivities': ['NetCashProvidedByUsedInInvestingActivities'],
            'PaymentsOfDividendsCommonStock': ['PaymentsOfDividendsCommonStock','PaymentsOfDividends'],
            'ProceedsFromIssuanceOrSaleOfEquity': ['ProceedsFromIssuanceOrSaleOfEquity','ProceedsFromStockOptionsExercised'],
            'PaymentsForRepurchaseOfCommonStock': ['PaymentsForRepurchaseOfCommonStock','PaymentsForRepurchaseOfEquity'],
            'PaymentsForExciseTaxOnPurchaseOfCommonStock': ['PaymentsForExciseTaxOnPurchaseOfCommonStock'],
            'ProceedsFromDebtMaturingInMoreThanThreeMonths': ['ProceedsFromDebtMaturingInMoreThanThreeMonths','ProceedsFromIssuanceOfLongTermDebt'],
            'RepaymentsOfDebtMaturingInMoreThanThreeMonths': ['RepaymentsOfDebtMaturingInMoreThanThreeMonths','RepaymentsOfLongTermDebt'],
            'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess': ['ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess','ProceedsFromRepaymentsOfShortTermDebt'],
            'NetCashProvidedByUsedInFinancingActivities': ['NetCashProvidedByUsedInFinancingActivities'],
            'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': ['EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents','EffectOfExchangeRateOnCashAndCashEquivalents'],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect': ['CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect','CashAndCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect'],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning': ['CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsBeginningBalance','CashAndCashEquivalentsAtCarryingValueBeginningBalance'],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End': ['CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents','CashAndCashEquivalentsAtCarryingValue']
        }

    def get_all_filings(self, start_date=None):
        """
        Get all filings from start_date to present
        Args:
            start_date: String in format 'YYYY-MM-DD' representing the earliest report date to include
        """
        url = f"{self.base_url}/submissions/CIK{self.cik}.json"
        logger.info(f"Fetching all filings since {start_date} for {self.company_name}")
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            time.sleep(0.1)
            data = response.json()
            recent = data['filings']['recent']
            filings = []
            for i in range(len(recent['form'])):
                form = recent['form'][i]
                if form in ['10-Q', '10-K']:
                    report_date = recent['reportDate'][i]
                    # Filter by report date, not filing date
                    if start_date is None or report_date >= start_date:
                        filings.append({
                            'accession': recent['accessionNumber'][i],
                            'filing_date': recent['filingDate'][i],
                            'report_date': report_date,
                            'form': form,
                            'primary_document': recent['primaryDocument'][i]
                        })
            filings.sort(key=lambda x: x['report_date'])
            logger.info(f"Found {len(filings)} filings since {start_date}")
            return filings
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching submissions: {e}")
            raise

    def construct_instance_url(self, accession, report_date):
        """Construct URL for XBRL instance document"""
        accession_no_dash = accession.replace('-', '')
        date_obj = datetime.strptime(report_date, '%Y-%m-%d')
        date_str = date_obj.strftime('%Y%m%d')
        url = f"{self.sec_archives}/{self.cik_int}/{accession_no_dash}/{self.ticker}-{date_str}_htm.xml"
        return url

    def download_instance_document(self, url):
        """Download XBRL instance document"""
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            time.sleep(0.15)
            return response.content
        except requests.exceptions.RequestException as e:
            logger.error(f"Error downloading {url}: {e}")
            raise

    def parse_context_elements(self, root):
        """Parse context elements to understand segments"""
        contexts = {}
        for context in root.findall('.//xbrli:context', self.namespaces):
            context_id = context.get('id')
            period = context.find('xbrli:period', self.namespaces)
            instant = period.find('xbrli:instant', self.namespaces)
            start = period.find('xbrli:startDate', self.namespaces)
            end = period.find('xbrli:endDate', self.namespaces)
            context_info = {
                'id': context_id,
                'segments': {},
                'instant': instant.text if instant is not None else None,
                'start': start.text if start is not None else None,
                'end': end.text if end is not None else None,
            }
            entity = context.find('xbrli:entity', self.namespaces)
            if entity is not None:
                segment = entity.find('xbrli:segment', self.namespaces)
                if segment is not None:
                    for member in segment.findall('.//xbrldi:explicitMember', self.namespaces):
                        dimension = member.get('dimension')
                        member_value = member.text
                        if ':' in member_value:
                            member_value = member_value.split(':')[1]
                        context_info['segments'][dimension] = member_value
            contexts[context_id] = context_info
        return contexts

    def extract_facts_from_xbrl(self, xml_content):
        """Extract all facts from XBRL instance, creating multiple entries for multi-dimensional contexts"""
        root = ET.fromstring(xml_content)
        # Update namespaces from document
        for prefix, uri in root.attrib.items():
            if prefix.startswith('{http://www.w3.org/2000/xmlns/}'):
                ns_prefix = prefix.split('}')[1]
                self.namespaces[ns_prefix] = uri
        contexts = self.parse_context_elements(root)
        facts = []
        for elem in root.iter():
            tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            context_ref = elem.get('contextRef')
            unit_ref = elem.get('unitRef')
            decimals = elem.get('decimals')
            if context_ref in contexts and elem.text:
                context = contexts[context_ref]
                
                try:
                    value = float(elem.text)
                except (ValueError, TypeError):
                    continue
                
                # NEW: Create facts for ALL relevant segment dimensions
                # This captures multi-dimensional contexts properly
                if context['segments']:
                    # Create a fact for each segment dimension
                    for dim, member in context['segments'].items():
                        fact = {
                            'tag': tag_name,
                            'value': value,
                            'context_id': context_ref,
                            'segment': member,
                            'dimension': dim,
                            'start_date': context['start'],
                            'end_date': context['end'],
                            'instant_date': context['instant'],
                            'decimals': decimals,
                            'unit': unit_ref
                        }
                        facts.append(fact)
                else:
                    # No segments - consolidated data
                    fact = {
                        'tag': tag_name,
                        'value': value,
                        'context_id': context_ref,
                        'segment': 'Consolidated',
                        'dimension': None,
                        'start_date': context['start'],
                        'end_date': context['end'],
                        'instant_date': context['instant'],
                        'decimals': decimals,
                        'unit': unit_ref
                    }
                    facts.append(fact)
        return facts

    def process_filing(self, filing):
        """Process a single filing"""
        try:
            url = self.construct_instance_url(filing['accession'], filing['report_date'])
            logger.info(f"Processing {filing['form']} from {filing['report_date']}")
            xml_content = self.download_instance_document(url)
            facts = self.extract_facts_from_xbrl(xml_content)
            for fact in facts:
                fact['accession'] = filing['accession']
                fact['filing_date'] = filing['filing_date']
                fact['report_date'] = filing['report_date']
                fact['form'] = filing['form']
            logger.info(f" Extracted {len(facts)} facts")
            return facts
        except Exception as e:
            logger.error(f"Error processing filing: {e}")
            return []

    def extract_all_data(self, start_date=None):
        """
        Extract all financial data
        Args:
            start_date: String in format 'YYYY-MM-DD' representing earliest report date to include
        """
        filings = self.get_all_filings(start_date=start_date)
        all_facts = []
        for i, filing in enumerate(filings, 1):
            logger.info(f"\n[{i}/{len(filings)}] " + "=" * 50)
            facts = self.process_filing(filing)
            all_facts.extend(facts)
        df = pd.DataFrame(all_facts)
        if not df.empty:
            for date_col in ['start_date', 'end_date', 'instant_date', 'filing_date', 'report_date']:
                if date_col in df.columns:
                    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            sort_col = 'end_date' if 'end_date' in df.columns else 'instant_date'
            df = df.sort_values([sort_col, 'tag', 'segment'], ascending=[True, True, True])
            logger.info(f"\nTotal facts extracted: {len(df)}")
        return df

    # --- Helpers for Q4 calculation ---
    def _is_calendar_ytd(self, row):
        """
        Determine if a 10-Q period is YTD for a calendar fiscal year (Jan 1 start).
        Adjust here if you later support non-calendar fiscal years.
        """
        s = row.get('start_date')
        e = row.get('end_date')
        if pd.isna(s) or pd.isna(e):
            return False
        return s.month == 1 and s.day == 1 and s.year == e.year

    def calculate_q4_data(self, df):
        """
        Calculate Q4 by subtracting YTD or available quarters from annual for flow statements,
        and copy 10-K data for balance sheet. **Consolidated only**.
        """
        logger.info("\n" + "=" * 60)
        logger.info("Calculating Q4 data (Consolidated only)")
        logger.info("=" * 60)
        if df.empty:
            return df

        quarterly_df = df[df['form'] == '10-Q'].copy()
        annual_df = df[df['form'] == '10-K'].copy()

        q4_records = []
        unique_tags = df['tag'].dropna().unique()
        # Only compute for Consolidated
        unique_segments = ['Consolidated']

        for tag in unique_tags:
            for segment in unique_segments:
                annual_subset = annual_df[
                    (annual_df['tag'] == tag) &
                    (
                        (annual_df['segment'] == segment) |
                        (segment == 'Consolidated' and annual_df['segment'].isna())
                    )
                ].copy()

                for _, annual_row in annual_subset.iterrows():
                    # Balance (instant) vs flow (period)
                    is_balance_sheet = pd.notna(annual_row.get('instant_date'))

                    if is_balance_sheet:
                        fiscal_year_end = annual_row['instant_date']
                        if pd.isna(fiscal_year_end):
                            continue
                        q4_records.append({
                            'tag': tag,
                            'value': annual_row['value'],
                            'segment': segment,
                            'start_date': None,
                            'end_date': None,
                            'instant_date': fiscal_year_end,
                            'report_date': fiscal_year_end,
                            'form': '10-K (Q4)',
                            'accession': annual_row['accession'],
                            'filing_date': annual_row['filing_date'],
                            'context_id': f"Q4_{fiscal_year_end.year}_{segment}",
                            'dimension': annual_row.get('dimension'),
                            'decimals': annual_row.get('decimals'),
                            'unit': annual_row.get('unit')
                        })
                        continue

                    # Flow statement (income, cash flow)
                    fiscal_year_end = annual_row['end_date']
                    if pd.isna(fiscal_year_end):
                        continue
                    fiscal_year = fiscal_year_end.year

                    q_subset = quarterly_df[
                        (quarterly_df['tag'] == tag) &
                        (
                            (quarterly_df['segment'] == segment) |
                            (segment == 'Consolidated' and quarterly_df['segment'].isna())
                        ) &
                        (quarterly_df['end_date'] > pd.Timestamp(year=fiscal_year - 1, month=12, day=31)) &
                        (quarterly_df['end_date'] <= fiscal_year_end)
                    ].copy()

                    if q_subset.empty:
                        continue

                    q_subset['is_ytd'] = q_subset.apply(self._is_calendar_ytd, axis=1)

                    annual_total = annual_row['value']
                    q3_end = q_subset['end_date'].max()

                    if q_subset['is_ytd'].any():
                        latest_ytd = q_subset.sort_values('end_date').iloc[-1]
                        q4_value = annual_total - latest_ytd['value']
                        logger.debug(
                            f"[Q4 YTD] tag={tag} segment={segment} FY={fiscal_year} "
                            f"Annual={annual_total} - Q3YTD={latest_ytd['value']} -> Q4={q4_value}"
                        )
                    else:
                        sum_quarters = q_subset['value'].sum()
                        q4_value = annual_total - sum_quarters
                        if len(q_subset) != 3:
                            logger.warning(
                                f"Discrete quarterly data incomplete for {tag} ({segment}) FY{fiscal_year}: "
                                f"have {len(q_subset)} quarters; Q4 computed as Annual - sum(available)"
                            )

                    q4_records.append({
                        'tag': tag,
                        'value': q4_value,
                        'segment': segment,
                        'start_date': q3_end + pd.Timedelta(days=1) if pd.notna(q3_end) else None,
                        'end_date': fiscal_year_end,
                        'instant_date': None,
                        'report_date': fiscal_year_end,
                        'form': '10-Q (Q4 Calculated)',
                        'accession': annual_row['accession'],
                        'filing_date': annual_row['filing_date'],
                        'context_id': f"Q4_{fiscal_year}_{segment}",
                        'dimension': annual_row.get('dimension'),
                        'decimals': annual_row.get('decimals'),
                        'unit': annual_row.get('unit')
                    })

        if q4_records:
            q4_df = pd.DataFrame(q4_records)
            combined_df = pd.concat([df, q4_df], ignore_index=True)
            # Keep chronological order by period/instant, then tag, then segment
            if 'end_date' in combined_df.columns:
                combined_df = combined_df.sort_values(['end_date', 'instant_date', 'tag', 'segment'])
            logger.info(f"Added {len(q4_records)} Q4 records (Consolidated only)")
            return combined_df

        return df

    def make_quarters_discrete(self, df):
        """
        Convert 10-Q YTD cash flow/income values to discrete quarter values by differencing.
        Applies to both Consolidated AND segment data. Skips '10-Q (Q4 Calculated)'.
        """
        if df.empty:
            return df

        # Real 10-Q rows, exclude calculated Q4
        is_real_q = (df['form'] == '10-Q') & df['end_date'].notna()
        q = df[is_real_q].copy()

        # Identify YTD rows
        q['is_ytd'] = q.apply(self._is_calendar_ytd, axis=1)

        out = []
        if not q.empty:
            # Group by tag, segment, and fiscal year, then apply differencing
            for (tag, segment, fy), g in q.groupby(['tag', 'segment', q['end_date'].dt.year]):
                g = g.sort_values('end_date')
                if g['is_ytd'].any():
                    vals = g['value'].values
                    g = g.copy()
                    # Q1_discrete = Q1_YTD; Q2_discrete = Q2_YTD - Q1_YTD; etc.
                    import numpy as np
                    g['value'] = np.diff(vals, prepend=0)
                out.append(g)
        
        if out:
            q_discrete = pd.concat(out)
            # Merge back: keep all non-YTD 10-Q rows and all other rows untouched
            df_non_ytd_10q = q[~q['is_ytd']].copy()
            df_other = df[~is_real_q]
            return pd.concat([df_other, df_non_ytd_10q, q_discrete], ignore_index=True)
        
        return df

    def create_statement_pivot(self, df, statement_type):
        """Create pivot table for a financial statement in proper order"""
        if df.empty:
            return pd.DataFrame()

        statement_items = self._get_statement_items(statement_type)

        # Determine date column
        if statement_type == 'balance':
            date_col = 'instant_date'
            df_filtered = df[df['instant_date'].notna()].copy()
        else:
            date_col = 'end_date'
            df_filtered = df[df['end_date'].notna()].copy()

        if df_filtered.empty:
            return pd.DataFrame()

        # Filter to quarterly data only (includes "10-Q (Q4 Calculated)")
        df_filtered = df_filtered[df_filtered['form'].str.contains('10-Q', na=False)]

        # If building a segment statement, exclude 10-K annual data and Q4 calculated
        segment_statement_types = {
            'constructionindustries', 'resourceindustries',
            'energyandtransportation', 'financialproducts', 'operatingsegments'
        }
        if statement_type in segment_statement_types:
            # Exclude 10-K annual reports (we only want quarterly 10-Q data)
            df_filtered = df_filtered[~df_filtered['form'].str.contains('10-K', na=False)]
            # Exclude Q4 calculated rows
            df_filtered = df_filtered[df_filtered['form'] != '10-Q (Q4 Calculated)']

        # Normalize segment text for safety (prevent None comparisons)
        df_filtered['segment'] = df_filtered['segment'].fillna('Consolidated')

        # Candidate tag maps
        income_tag_map = self._get_income_tag_candidates() if statement_type == 'income' else {}
        cashflow_tag_map = self._get_cashflow_tag_candidates() if statement_type == 'cashflow' else {}
        balance_tag_map = self._get_balance_tag_candidates() if statement_type == 'balance' else {}

        # Map segment suffix to expected member names
        segment_map = {
            'FP': 'FinancialProductsMember',
            'MET': 'MachineryEnergyTransportationMember',
            'EXFP': 'AllOtherExcludingFinancialProductsMember',
            'Total': 'Consolidated',
            'OpSeg': 'OperatingSegmentsMember',  # Aggregate of all operating segments
            'CI': 'ConstructionIndustriesMember',
            'RI': 'ResourceIndustriesMember',
            'ET': 'EnergyandTransportationMember',  # Note: lowercase "and"
            'FPS': 'FinancialProductsSegmentMember'
        }

        pivot_data = []
        for tag_key, label in statement_items.items():
            if tag_key == '':
                pivot_data.append({'Line_Item': label})
                continue

            # Extract base tag and optional segment suffix from our key
            parts = tag_key.split('_', 1)
            base_key = parts[0]
            segment_suffix = parts[1] if len(parts) > 1 else None

            # Determine candidate tags
            if statement_type == 'income':
                candidate_tags = income_tag_map.get(base_key, [base_key])
            elif statement_type == 'cashflow':
                candidate_tags = cashflow_tag_map.get(base_key, [base_key])
            elif statement_type == 'balance':
                candidate_tags = balance_tag_map.get(base_key, [base_key])
            else:
                candidate_tags = [base_key if segment_suffix else tag_key] if segment_suffix else [tag_key]

            # Build subset using segment targeting
            if segment_suffix:
                target_segment = segment_map.get(segment_suffix, segment_suffix)
                # Try exact match first
                subset = df_filtered[
                    (df_filtered['tag'].isin(candidate_tags)) &
                    (df_filtered['segment'] == target_segment)
                ]
                # Fallback: case-insensitive contains (to catch slight naming variants)
                if subset.empty:
                    subset = df_filtered[
                        (df_filtered['tag'].isin(candidate_tags)) &
                        (df_filtered['segment'].str.contains(str(target_segment), case=False, na=False))
                    ]
                # Last-resort: weak keyword match on tag
                if subset.empty:
                    subset = df_filtered[
                        df_filtered['tag'].str.contains(base_key, case=False, regex=False, na=False) &
                        (df_filtered['segment'] == target_segment)
                    ]
            else:
                # Consolidated / no segment
                subset = df_filtered[
                    (df_filtered['tag'].isin(candidate_tags)) &
                    (df_filtered['segment'].isin(['Consolidated', '']))
                ]
                if subset.empty:
                    subset = df_filtered[
                        df_filtered['tag'].str.contains(base_key, case=False, regex=False, na=False) &
                        (df_filtered['segment'].isin(['Consolidated', '']))
                    ]

            if not subset.empty:
                # Deduplicate: For multi-dimensional contexts, we may have duplicate rows
                # Prioritize the correct dimension based on segment type
                subset = subset.copy()
                
                # Individual segments use StatementBusinessSegmentsAxis
                # Operating Segments aggregate uses ConsolidationItemsAxis
                if statement_type in {'constructionindustries', 'resourceindustries', 
                                     'energyandtransportation', 'financialproducts'}:
                    # Prioritize StatementBusinessSegmentsAxis for individual segments
                    subset['priority'] = subset['dimension'].apply(
                        lambda x: 0 if 'StatementBusinessSegmentsAxis' in str(x) else 1
                    )
                elif statement_type == 'operatingsegments':
                    # Prioritize ConsolidationItemsAxis for aggregate operating segments
                    subset['priority'] = subset['dimension'].apply(
                        lambda x: 0 if 'ConsolidationItemsAxis' in str(x) else 1
                    )
                else:
                    subset['priority'] = 1
                
                subset = subset.sort_values('priority')
                subset = subset.drop(columns=['priority'])
                
                # Drop duplicates keeping first (highest priority)
                subset = subset.drop_duplicates(subset=['tag', date_col], keep='first')
                
                pivot_row = subset.pivot_table(
                    index='tag',
                    columns=date_col,
                    values='value',
                    aggfunc='first'
                )
                if not pivot_row.empty:
                    row_dict = {'Line_Item': label}
                    row_dict.update(pivot_row.iloc[0].to_dict())
                    pivot_data.append(row_dict)

        if not pivot_data:
            return pd.DataFrame()

        result_df = pd.DataFrame(pivot_data)
        # Sort columns by date (most recent first), keeping Line_Item first
        date_columns = [col for col in result_df.columns if col != 'Line_Item']
        # Convert any pandas Timestamps to strings 'YYYY-MM-DD' for ordering & output
        normalized_cols = []
        for col in date_columns:
            if isinstance(col, pd.Timestamp):
                normalized_cols.append(col)
            else:
                try:
                    normalized_cols.append(pd.to_datetime(col))
                except Exception:
                    normalized_cols.append(col)
        # Reorder
        try:
            order = list(sorted([c for c in normalized_cols if isinstance(c, pd.Timestamp)], reverse=True))
        except Exception:
            order = date_columns
        # Build final columns
        final_cols = ['Line_Item'] + order
        result_df = result_df.reindex(columns=final_cols)

        # Format column names that are datetimes to YYYY-MM-DD
        formatted_cols = []
        for col in result_df.columns:
            if isinstance(col, pd.Timestamp):
                formatted_cols.append(col.strftime('%Y-%m-%d'))
            else:
                formatted_cols.append(col)
        result_df.columns = formatted_cols
        return result_df

    def _get_quarter_from_date(self, date_str):
        """Determine quarter from date string"""
        try:
            if isinstance(date_str, str):
                date_obj = pd.to_datetime(date_str)
            else:
                date_obj = date_str
            month = date_obj.month
            # Calendar fiscal year (Dec year-end): Q1 Jan-Mar, Q2 Apr-Jun, Q3 Jul-Sep, Q4 Oct-Dec
            if month in [1, 2, 3]:
                return 'Q1'
            elif month in [4, 5, 6]:
                return 'Q2'
            elif month in [7, 8, 9]:
                return 'Q3'
            elif month in [10, 11, 12]:
                return 'Q4'
            else:
                return ''
        except Exception:
            return ''

    def format_excel_sheet(self, writer, sheet_name, df):
        """Apply formatting to Excel sheet with quarter labels"""
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Insert a new row at the top for quarter labels
        worksheet.insert_rows(1)

        # Add quarter labels for each date column
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name != 'Line_Item' and isinstance(col_name, str):
                quarter = self._get_quarter_from_date(col_name)
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = quarter

        # Format the quarter row
        quarter_format = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        quarter_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in worksheet[1]:
            cell.fill = quarter_format
            cell.font = quarter_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Format the date header row (now row 2)
        header_format = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in worksheet[2]:
            cell.fill = header_format
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    text = str(cell.value) if cell.value is not None else ""
                    if len(text) > max_length:
                        max_length = len(text)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Format number cells (starting from row 3 now)
        accounting_format = '#,##0'
        for row in worksheet.iter_rows(min_row=3):
            for cell in row[1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = accounting_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')

        # Freeze panes to keep quarter and date headers visible
        worksheet.freeze_panes = 'B3'

    def export_to_excel(self, output_filename, start_date='2020-01-01'):
        """
        Extract and export all financial data
        Args:
            output_filename: Name of Excel file to create
            start_date: String in format 'YYYY-MM-DD' for earliest report date to include
        """
        logger.info("=" * 60)
        logger.info(f"Starting comprehensive extraction for {self.company_name}")
        logger.info(f"Data range: {start_date} - Present")
        logger.info("=" * 60)

        df = self.extract_all_data(start_date)
        if df.empty:
            logger.warning("No data extracted!")
            return None

        df = self.calculate_q4_data(df)
        df = self.make_quarters_discrete(df)

        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Raw data (reported + consolidated calculated rows)
            df.to_excel(writer, sheet_name='All Data - Raw', index=False)
            self.format_excel_sheet(writer, 'All Data - Raw', df)

            # Income Statement
            logger.info("Creating Income Statement")
            income_pivot = self.create_statement_pivot(df, 'income')
            if not income_pivot.empty:
                income_pivot.to_excel(writer, sheet_name='Income Statement - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Income Statement - Quarterly', income_pivot)

            # Balance Sheet
            logger.info("Creating Balance Sheet")
            balance_pivot = self.create_statement_pivot(df, 'balance')
            if not balance_pivot.empty:
                balance_pivot.to_excel(writer, sheet_name='Balance Sheet - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Balance Sheet - Quarterly', balance_pivot)

            # Cash Flow
            logger.info("Creating Cash Flow Statement")
            cashflow_pivot = self.create_statement_pivot(df, 'cashflow')
            if not cashflow_pivot.empty:
                cashflow_pivot.to_excel(writer, sheet_name='Cash Flow - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Cash Flow - Quarterly', cashflow_pivot)

            # Operating Segments (Aggregate) - Has COGS/SG&A starting in 2024
            logger.info("Creating Operating Segments (Aggregate) Statement")
            operatingsegments_pivot = self.create_statement_pivot(df, 'operatingsegments')
            if not operatingsegments_pivot.empty:
                operatingsegments_pivot.to_excel(writer, sheet_name='Operating Segments - Qtrly', index=False)
                self.format_excel_sheet(writer, 'Operating Segments - Qtrly', operatingsegments_pivot)

            # Construction Industries Segment
            logger.info("Creating Construction Industries Segment Statement")
            constructionindustries_pivot = self.create_statement_pivot(df, 'constructionindustries')
            if not constructionindustries_pivot.empty:
                constructionindustries_pivot.to_excel(writer, sheet_name='Construction Industries', index=False)
                self.format_excel_sheet(writer, 'Construction Industries', constructionindustries_pivot)

            # Resource Industries Segment
            logger.info("Creating Resource Industries Segment Statement")
            resourceindustries_pivot = self.create_statement_pivot(df, 'resourceindustries')
            if not resourceindustries_pivot.empty:
                resourceindustries_pivot.to_excel(writer, sheet_name='Resource Industries - Qtrly', index=False)
                self.format_excel_sheet(writer, 'Resource Industries - Qtrly', resourceindustries_pivot)

            # Energy & Transportation Segment
            logger.info("Creating Energy & Transportation Segment Statement")
            energyandtransportation_pivot = self.create_statement_pivot(df, 'energyandtransportation')
            if not energyandtransportation_pivot.empty:
                energyandtransportation_pivot.to_excel(writer, sheet_name='Energy & Transport - Qtrly', index=False)
                self.format_excel_sheet(writer, 'Energy & Transport - Qtrly', energyandtransportation_pivot)

            # Financial Products Segment
            logger.info("Creating Financial Products Segment Statement")
            financialproducts_pivot = self.create_statement_pivot(df, 'financialproducts')
            if not financialproducts_pivot.empty:
                financialproducts_pivot.to_excel(writer, sheet_name='Financial Products - Qtrly', index=False)
                self.format_excel_sheet(writer, 'Financial Products - Qtrly', financialproducts_pivot)

        logger.info("=" * 60)
        logger.info(f"Export complete! File saved: {output_filename}")
        logger.info("=" * 60)
        return output_filename


def main():
    """Main execution"""
    YOUR_EMAIL = "brayden.joyce@doosan.com"
    CIK = "0000018230"
    COMPANY_NAME = "Caterpillar Inc."
    TICKER = "cat"
    START_DATE = "2020-03-31"  # Q1 2020 report date

    extractor = ComprehensiveXBRLExtractor(
        email=YOUR_EMAIL,
        cik=CIK,
        company_name=COMPANY_NAME,
        ticker=TICKER
    )

    output_file = extractor.export_to_excel(
        output_filename='caterpillar_financials.xlsx',
        start_date=START_DATE
    )

    if output_file:
        print(f"\nSuccess! Complete financial data exported to: {output_file}")
        print(f"\nData range: {START_DATE} - Present")
        print("\nThe Excel file contains:")
        print(" Income Statement - Quarterly")
        print(" Balance Sheet - Quarterly")
        print(" Cash Flow Statement - Quarterly")
        print(" Construction Industries Segment - Quarterly")
        print(" Resource Industries Segment - Quarterly")
        print(" Energy & Transportation Segment - Quarterly")
        print(" Financial Products Segment - Quarterly")
        print(" All raw data")


if __name__ == "__main__":
    main()
