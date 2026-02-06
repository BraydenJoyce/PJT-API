"""
SEC EDGAR XBRL Parser - Complete Financial Statement Extractor
Extracts both consolidated and segment-level data with Q4 calculations

MERGED VERSION - Best of Both Worlds:
- Robust EDGAR instance discovery (index.json + HTML fallback) - no more 404 errors
- Comprehensive tag catalogs for all financial statements
- YTD-aware Q4 calculations with discrete quarterly cash flow normalization
- Professional Excel formatting with quarter labels

Author: Brayden Joyce
Company: Doosan/Bobcat Company
"""

import re
import time
import json
import logging
from datetime import datetime
from collections import OrderedDict

import requests
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
# Logging Configuration
# -----------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ComprehensiveXBRLExtractor:
    """Extract complete financial statements with segment breakdowns"""

    def __init__(self, email, cik, company_name, ticker):
        """
        Initialize the extractor
        
        Args:
            email: Your email for SEC API user agent (required by SEC)
            cik: Company CIK number (with leading zeros, e.g., "0000018230")
            company_name: Company name for logging
            ticker: Company ticker symbol (lowercase, for file identification)
        """
        self.base_url = "https://data.sec.gov"
        self.sec_archives = "https://www.sec.gov/Archives/edgar/data"
        self.headers = {
            # SEC fair access: identifiable User-Agent with email. 10 req/s max.
            # https://www.sec.gov/search-filings/edgar-search-assistance/accessing-edgar-data
            'User-Agent': f'{email}',
            'Accept-Encoding': 'gzip, deflate'
        }
        self.cik = cik  # "0000018230"
        self.cik_int = str(int(cik))  # "18230"
        self.company_name = company_name
        self.ticker = ticker.lower()

        # XBRL namespaces (updated from document during parsing)
        self.namespaces = {
            'xbrli': 'http://www.xbrl.org/2003/instance',
            'xbrldi': 'http://xbrl.org/2006/xbrldi',
        }

    # -------------------------------------------------------------------------
    # Statement Layout Definitions
    # -------------------------------------------------------------------------
    def _get_statement_items(self, statement_type):
        """Get line items in proper order for each statement type"""
        
        if statement_type == 'income':
            return OrderedDict([
                # Sales and Revenues
                ('Revenues_MET', '     Sales of Machinery, Energy & Transportation'),
                ('Revenues_FinancialProducts', '     Revenues of Financial Products'),
                ('Revenues_Total', '     Total sales and revenues'),
                
                # Operating Costs
                ('CostOfRevenue', '     Cost of goods sold'),
                ('SellingGeneralAndAdministrativeExpense', '     SG&A Expenses'),
                ('ResearchAndDevelopmentExpense', '     R&D Expenses'),
                ('FinancingInterestExpense_FinancialProducts', '     Interest expense of Financial Products'),
                ('OtherOperatingIncomeExpenseNet', '     Other operating (income) expenses'),
                ('CostsAndExpenses', '     Total operating costs'),
                
                # Operating Profit and Below
                ('OperatingIncomeLoss', 'Operating Profit'),
                ('InterestExpenseNonoperating_EXFP', '     Interest expense excluding Financial Products'),
                ('OtherNonoperatingIncomeExpense', '     Other income (expense)'),
                ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'Consolidated profit before taxes'),
                ('IncomeTaxExpenseBenefit', '     Provision (benefit) for income taxes'),
                ('ProfitOfConsolidatedCompanies', '     Profit of consolidated companies'),
                ('IncomeLossFromEquityMethodInvestments', '     Equity in profit (loss) of unconsolidated affiliated companies'),
                ('ProfitLoss', 'Profit of consolidated and affiliated companies'),
                ('NetIncomeLossAttributableToNoncontrollingInterest', 'Profit (loss) attributable to noncontrolling interests'),
                ('NetIncomeLossAvailableToCommonStockholdersBasic', 'Profit (Attributable to Common Stockholders)'),
                
                # EPS and Shares
                ('EarningsPerShareBasic', 'Profit per common share'),
                ('EarningsPerShareDiluted', 'Profit per common share - diluted'),
                ('WeightedAverageNumberOfSharesOutstandingBasic', '     Shares Outstanding - Basic'),
                ('WeightedAverageNumberOfDilutedSharesOutstanding', '     Shares Outstanding - Diluted'),
            ])
            
        elif statement_type == 'balance':
            return OrderedDict([
                # Current Assets
                ('CashAndCashEquivalentsAtCarryingValue', '          Cash & Cash Equivalents'),
                ('AccountsReceivableNetCurrent', '          Receivables - trade and other'),
                ('NotesAndLoansReceivableNetCurrent', '          Receivables - finance'),
                ('PrepaidExpenseAndOtherAssetsCurrent', '          Prepaid Expenses And Other Assets Current'),
                ('InventoryNet', '          Inventories'),
                ('AssetsCurrent', '     Total Current Assets'),
                
                # Noncurrent Assets
                ('PropertyPlantAndEquipmentNet', '     Property, Plant, & Equipment - net'),
                ('AccountsReceivableNetNoncurrent', '     Long-term receivables - trade and other'),
                ('NotesAndLoansReceivableNetNoncurrent', '     Long-term receivables - finance'),
                ('NoncurrentDeferredAndRefundableIncomeTaxes', '     Noncurrent deferred and refundable income taxes'),
                ('IntangibleAssetsNetExcludingGoodwill', '     Intangible Assets'),
                ('Goodwill', '     Goodwill'),
                ('OtherAssetsNoncurrent', '     Other assets'),
                ('Assets', 'Total assets'),
                
                # Current Liabilities
                ('ShortTermBorrowings_FinancialProducts', '          Short Term Borrowings: Financial Products'),
                ('AccountsPayableCurrent', '          Accounts payable'),
                ('AccruedLiabilitiesCurrent', '          Accrued expenses'),
                ('EmployeeRelatedLiabilitiesCurrent', '          Accrued wages: salaries and employee benefits'),
                ('ContractWithCustomerLiabilityCurrent', '          Customer advances'),
                ('DividendsPayableCurrent', '          Dividends payable'),
                ('OtherLiabilitiesCurrent', '          Other current liabilities'),
                ('LongTermDebtAndCapitalLeaseObligationsCurrent_MET', '               Long-term debt due within one year: Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligationsCurrent_FinancialProducts', '               Long-term debt due within one year: Financial Products'),
                ('LiabilitiesCurrent', '     Total current liabilities'),
                
                # Noncurrent Liabilities
                ('LongTermDebtAndCapitalLeaseObligations_MET', '               Long-term debt due after one year: Machinery: Energy & Transportation'),
                ('LongTermDebtAndCapitalLeaseObligations_FinancialProducts', '               Long-term debt due after one year: Financial Products'),
                ('PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent', '     Liability for postemployment benefits'),
                ('OtherLiabilitiesNoncurrent', '     Other liabilities'),
                ('Liabilities', 'Total Liabilities'),
                
                # Shareholders' Equity
                ('CommonStocksIncludingAdditionalPaidInCapital', '     Issued shares at paid-in amount'),
                ('TreasuryStockValue', '     Treasury stock at cost'),
                ('RetainedEarningsAccumulatedDeficit', '     Profit employed in the business'),
                ('AccumulatedOtherComprehensiveIncomeLossNetOfTax', '     Accumulated other comprehensive income (loss)'),
                ('MinorityInterest', '     Noncontrolling interests'),
                ('StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', "Total shareholders' equity"),
                ('LiabilitiesAndStockholdersEquity', "Total liabilities and shareholders' equity"),
            ])
            
        elif statement_type == 'cashflow':
            return OrderedDict([
                # Operating Activities
                ('ProfitLoss', '     Profit of consolidated and affiliated companies'),
                
                # Adjustments to reconcile profit
                ('DepreciationDepletionAndAmortization', '          Depreciation and Amortization'),
                ('DeferredIncomeTaxExpenseBenefit', '          Provision (benefit) for deferred income taxes'),
                ('NonCashGainLossOnDivestiture', '          (Gain) loss on divestiture'),
                ('OtherNoncashIncomeExpense', '          Other'),
                
                # Changes in assets and liabilities
                ('IncreaseDecreaseInReceivables', '          Receivables – trade and other'),
                ('IncreaseDecreaseInInventories', '          Inventories'),
                ('IncreaseDecreaseInAccountsPayable', '          Accounts payable'),
                ('IncreaseDecreaseInAccruedLiabilities', '          Accrued expenses'),
                ('IncreaseDecreaseInEmployeeRelatedLiabilities', '          Accrued wages, salaries and employee benefits'),
                ('IncreaseDecreaseInContractWithCustomerLiability', '          Customer advances'),
                ('IncreaseDecreaseInOtherOperatingAssets', '          Other assets - net'),
                ('IncreaseDecreaseInOtherOperatingLiabilities', '          Other liabilities - net'),
                ('NetCashProvidedByUsedInOperatingActivities', 'Net cash provided by (used for) operating activities'),
                
                # Investing Activities
                ('PaymentsToAcquirePropertyPlantAndEquipment', '     Capital expenditures – excluding equipment leased to others'),
                ('PaymentsToAcquireEquipmentOnLease', '     Expenditures for equipment leased to others'),
                ('ProceedsFromSaleOfPropertyPlantAndEquipment', '     Proceeds from disposals of leased assets and property, plant and equipment'),
                ('PaymentsToAcquireFinanceReceivables', '     Additions to finance receivables'),
                ('ProceedsFromCollectionOfFinanceReceivables', '     Collections of finance receivables'),
                ('ProceedsFromSaleOfFinanceReceivables', '     Proceeds from sale of finance receivables'),
                ('PaymentsToAcquireBusinessesNetOfCashAcquired', '     Investments and acquisitions (net of cash acquired)'),
                ('ProceedsFromDivestitureOfBusinessesNetOfCashDivested', '     Proceeds from sale of businesses and investments (net of cash sold)'),
                ('ProceedsFromSaleAndMaturityOfMarketableSecurities', '     Proceeds from maturities and sale of securities'),
                ('PaymentsToAcquireMarketableSecurities', '     Investments in securities'),
                ('PaymentsForProceedsFromOtherInvestingActivities', '     Other – net'),
                ('NetCashProvidedByUsedInInvestingActivities', 'Net cash provided by (used for) investing activities'),
                
                # Financing Activities
                ('PaymentsOfDividendsCommonStock', '     Dividends paid'),
                ('ProceedsFromIssuanceOrSaleOfEquity', '     Common stock issued, and other stock compensation transactions, net'),
                ('PaymentsForRepurchaseOfCommonStock', '     Payments to purchase common stock'),
                ('PaymentsForExciseTaxOnPurchaseOfCommonStock', '     Excise tax paid on purchases of common stock'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_MET', '     Proceeds from debt issued (original maturities greater than three months) - Machinery, Energy & Transportation'),
                ('ProceedsFromDebtMaturingInMoreThanThreeMonths_FinancialProducts', '     Proceeds from debt issued (original maturities greater than three months) - Financial Products'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_MET', '     Payments on debt (original maturities greater than three months) - Machinery, Energy & Transportation'),
                ('RepaymentsOfDebtMaturingInMoreThanThreeMonths_FinancialProducts', '     Payments on debt (original maturities greater than three months) - Financial Products'),
                ('ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess', '     Short-term borrowings – net (original maturities three months or less)'),
                ('NetCashProvidedByUsedInFinancingActivities', 'Net cash provided by (used for) financing activities'),
                
                # Cash reconciliation
                ('EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents', 'Effect of exchange rate changes on cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect', 'Increase (decrease) in cash, cash equivalents and restricted cash'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning', 'Cash, cash equivalents and restricted cash at beginning of period'),
                ('CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End', 'Cash, cash equivalents and restricted cash at end of period'),
            ])
        
        logger.warning(f"Unknown statement type: {statement_type}")
        return OrderedDict()

    def _get_segment_items(self):
        """Get line items for individual segment sheets"""
        return OrderedDict([
            ('Revenues', 'Sales and revenues'),
            ('CostOfRevenue', 'Cost of goods sold'),
            ('SellingGeneralAndAdministrativeResearchAndDevelopment', 'SG&A/R&D'),
            ('SegmentReportingOtherItemAmount', 'Other segment items'),
            ('IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments', 'Segment Profit'),
            ('Assets', 'Assets'),
            ('DepreciationDepletionAndAmortization', 'Depreciation and amortization'),
            ('SegmentReportingInformationExpenditureForAdditionsLongLivedAssets', 'Capital expenditures'),
        ])

    # -------------------------------------------------------------------------
    # Candidate Tag Catalogs (for robust matching)
    # -------------------------------------------------------------------------
    def _get_income_tag_candidates(self):
        """Comprehensive tag candidates for income statement items"""
        return {
            'Revenues': [
                'Revenues', 'SalesRevenueNet', 'SalesAndRevenue', 'SalesRevenueGoodsNet',
                'SalesRevenueServicesNet', 'RevenueFromContractWithCustomerExcludingAssessedTax'
            ],
            'CostOfRevenue': ['CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfSales'],
            'SellingGeneralAndAdministrativeExpense': ['SellingGeneralAndAdministrativeExpense'],
            'ResearchAndDevelopmentExpense': ['ResearchAndDevelopmentExpense'],
            'OtherOperatingIncomeExpenseNet': ['OtherOperatingIncomeExpenseNet', 'OtherOperatingIncomeExpense'],
            'CostsAndExpenses': ['CostsAndExpenses', 'OperatingExpenses'],
            'OperatingIncomeLoss': ['OperatingIncomeLoss'],
            'InterestExpenseNonoperating': ['InterestExpenseNonoperating', 'InterestAndDebtExpense'],
            'InterestExpenseNonoperating_EXFP': ['InterestExpenseExcludingFinancialProducts'],
            'OtherNonoperatingIncomeExpense': ['OtherNonoperatingIncomeExpense', 'NonoperatingIncomeExpense'],
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': [
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments',
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxes'
            ],
            'IncomeTaxExpenseBenefit': ['IncomeTaxExpenseBenefit'],
            'IncomeLossFromEquityMethodInvestments': ['IncomeLossFromEquityMethodInvestments'],
            'ProfitLoss': ['ProfitLoss', 'NetIncomeLoss'],
            'NetIncomeLossAttributableToNoncontrollingInterest': ['NetIncomeLossAttributableToNoncontrollingInterest'],
            'NetIncomeLossAvailableToCommonStockholdersBasic': ['NetIncomeLossAvailableToCommonStockholdersBasic'],
            'EarningsPerShareBasic': ['EarningsPerShareBasic'],
            'EarningsPerShareDiluted': ['EarningsPerShareDiluted'],
            'WeightedAverageNumberOfSharesOutstandingBasic': ['WeightedAverageNumberOfSharesOutstandingBasic'],
            'WeightedAverageNumberOfDilutedSharesOutstanding': ['WeightedAverageNumberOfDilutedSharesOutstanding'],
            'ProfitOfConsolidatedCompanies': ['ProfitOfConsolidatedCompanies', 'IncomeLossFromContinuingOperationsAfterIncomeTaxes'],
            'FinancingInterestExpense_FinancialProducts': ['InterestExpense', 'InterestAndDebtExpense'],
        }

    def _get_balance_tag_candidates(self):
        """Comprehensive tag candidates for balance sheet items"""
        return {
            'CashAndCashEquivalentsAtCarryingValue': [
                'CashAndCashEquivalentsAtCarryingValue', 'CashAndCashEquivalents'
            ],
            'AccountsReceivableNetCurrent': [
                'AccountsReceivableNetCurrent', 'ReceivablesNetCurrent', 'TradeAccountsReceivableNetCurrent'
            ],
            'NotesAndLoansReceivableNetCurrent': [
                'NotesAndLoansReceivableNetCurrent', 'LoansReceivableNetCurrent', 'FinanceReceivablesNetCurrent'
            ],
            'PrepaidExpenseAndOtherAssetsCurrent': [
                'PrepaidExpenseAndOtherAssetsCurrent', 'PrepaidExpenseCurrent', 'PrepaidExpenseAndOtherCurrentAssets'
            ],
            'InventoryNet': ['InventoryNet', 'InventoriesNet', 'InventoryAndOtherCurrentAssets'],
            'AssetsCurrent': ['AssetsCurrent', 'CurrentAssets'],
            
            'PropertyPlantAndEquipmentNet': [
                'PropertyPlantAndEquipmentNet', 'PropertyPlantAndEquipmentExcludingLandNet'
            ],
            'AccountsReceivableNetNoncurrent': ['AccountsReceivableNetNoncurrent', 'ReceivablesNetNoncurrent'],
            'NotesAndLoansReceivableNetNoncurrent': [
                'NotesAndLoansReceivableNetNoncurrent', 'LoansReceivableNetNoncurrent', 'FinanceReceivablesNetNoncurrent'
            ],
            'NoncurrentDeferredAndRefundableIncomeTaxes': [
                'NoncurrentDeferredAndRefundableIncomeTaxes', 'DeferredTaxAssetsNetNoncurrent', 
                'DeferredIncomeTaxAssetsNetNoncurrent'
            ],
            'IntangibleAssetsNetExcludingGoodwill': ['IntangibleAssetsNetExcludingGoodwill', 'IntangibleAssetsNet'],
            'Goodwill': ['Goodwill'],
            'OtherAssetsNoncurrent': ['OtherAssetsNoncurrent'],
            'Assets': ['Assets', 'AssetsNet'],
            
            'ShortTermBorrowings': ['ShortTermBorrowings', 'CommercialPaper', 'ShortTermDebt'],
            'AccountsPayableCurrent': ['AccountsPayableCurrent'],
            'AccruedLiabilitiesCurrent': ['AccruedLiabilitiesCurrent', 'OtherAccruedLiabilitiesCurrent'],
            'EmployeeRelatedLiabilitiesCurrent': ['EmployeeRelatedLiabilitiesCurrent', 'EmployeeRelatedLiabilities'],
            'ContractWithCustomerLiabilityCurrent': ['ContractWithCustomerLiabilityCurrent', 'DeferredRevenueCurrent'],
            'DividendsPayableCurrent': ['DividendsPayableCurrent', 'DividendsPayable'],
            'OtherLiabilitiesCurrent': ['OtherLiabilitiesCurrent'],
            'LiabilitiesCurrent': ['LiabilitiesCurrent', 'CurrentLiabilities'],
            
            'LongTermDebtAndCapitalLeaseObligationsCurrent': [
                'LongTermDebtAndCapitalLeaseObligationsCurrent', 'LongTermDebtCurrent'
            ],
            'LongTermDebtAndCapitalLeaseObligations': [
                'LongTermDebtAndCapitalLeaseObligations', 'LongTermDebtNoncurrent', 'LongTermDebt'
            ],
            'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent': [
                'PensionAndOtherPostretirementAndPostemploymentBenefitPlansLiabilitiesNoncurrent',
                'PensionAndOtherPostretirementDefinedBenefitPlansLiabilityNoncurrent'
            ],
            'OtherLiabilitiesNoncurrent': ['OtherLiabilitiesNoncurrent'],
            'Liabilities': ['Liabilities'],
            
            'CommonStocksIncludingAdditionalPaidInCapital': [
                'CommonStocksIncludingAdditionalPaidInCapital', 'CommonStockValue', 'StockIssuedDuringPeriodValueNewIssues'
            ],
            'TreasuryStockValue': ['TreasuryStockValue'],
            'RetainedEarningsAccumulatedDeficit': ['RetainedEarningsAccumulatedDeficit', 'RetainedEarnings'],
            'AccumulatedOtherComprehensiveIncomeLossNetOfTax': [
                'AccumulatedOtherComprehensiveIncomeLossNetOfTax', 'AccumulatedOtherComprehensiveIncomeLoss'
            ],
            'MinorityInterest': ['MinorityInterest', 'StockholdersEquityAttributableToNoncontrollingInterest'],
            'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest': [
                'StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest', 
                'StockholdersEquity'
            ],
            'LiabilitiesAndStockholdersEquity': ['LiabilitiesAndStockholdersEquity'],
        }

    def _get_cashflow_tag_candidates(self):
        """Comprehensive tag candidates for cash flow statement items"""
        return {
            'ProfitLoss': ['ProfitLoss', 'NetIncomeLoss'],
            'DepreciationDepletionAndAmortization': [
                'DepreciationDepletionAndAmortization', 'DepreciationAndAmortization'
            ],
            'DeferredIncomeTaxExpenseBenefit': ['DeferredIncomeTaxExpenseBenefit', 'DeferredIncomeTaxes'],
            'NonCashGainLossOnDivestiture': ['NonCashGainLossOnDivestiture', 'GainLossOnSaleOfBusiness'],
            'OtherNoncashIncomeExpense': ['OtherNoncashIncomeExpense', 'OtherNoncashIncome'],
            
            'IncreaseDecreaseInReceivables': [
                'IncreaseDecreaseInReceivables', 'IncreaseDecreaseInAccountsReceivable'
            ],
            'IncreaseDecreaseInInventories': ['IncreaseDecreaseInInventories'],
            'IncreaseDecreaseInAccountsPayable': ['IncreaseDecreaseInAccountsPayable'],
            'IncreaseDecreaseInAccruedLiabilities': [
                'IncreaseDecreaseInAccruedLiabilities', 'IncreaseDecreaseInOtherAccruedLiabilities'
            ],
            'IncreaseDecreaseInEmployeeRelatedLiabilities': ['IncreaseDecreaseInEmployeeRelatedLiabilities'],
            'IncreaseDecreaseInContractWithCustomerLiability': [
                'IncreaseDecreaseInContractWithCustomerLiability', 'IncreaseDecreaseInDeferredRevenue'
            ],
            'IncreaseDecreaseInOtherOperatingAssets': [
                'IncreaseDecreaseInOtherOperatingAssets', 'IncreaseDecreaseInOtherAssets'
            ],
            'IncreaseDecreaseInOtherOperatingLiabilities': [
                'IncreaseDecreaseInOtherOperatingLiabilities', 'IncreaseDecreaseInOtherLiabilities'
            ],
            'NetCashProvidedByUsedInOperatingActivities': [
                'NetCashProvidedByUsedInOperatingActivities',
                'NetCashProvidedByUsedInOperatingActivitiesContinuingOperations'
            ],
            
            'PaymentsToAcquirePropertyPlantAndEquipment': [
                'PaymentsToAcquirePropertyPlantAndEquipment', 'CapitalExpenditures'
            ],
            'PaymentsToAcquireEquipmentOnLease': ['PaymentsToAcquireEquipmentOnLease'],
            'ProceedsFromSaleOfPropertyPlantAndEquipment': ['ProceedsFromSaleOfPropertyPlantAndEquipment'],
            'PaymentsToAcquireFinanceReceivables': ['PaymentsToAcquireFinanceReceivables'],
            'ProceedsFromCollectionOfFinanceReceivables': ['ProceedsFromCollectionOfFinanceReceivables'],
            'ProceedsFromSaleOfFinanceReceivables': ['ProceedsFromSaleOfFinanceReceivables'],
            'PaymentsToAcquireBusinessesNetOfCashAcquired': [
                'PaymentsToAcquireBusinessesNetOfCashAcquired', 'PaymentsForAcquisitions'
            ],
            'ProceedsFromDivestitureOfBusinessesNetOfCashDivested': [
                'ProceedsFromDivestitureOfBusinessesNetOfCashDivested', 'ProceedsFromSaleOfBusiness'
            ],
            'ProceedsFromSaleAndMaturityOfMarketableSecurities': [
                'ProceedsFromSaleAndMaturityOfMarketableSecurities', 
                'ProceedsFromSaleOfMarketableSecurities'
            ],
            'PaymentsToAcquireMarketableSecurities': ['PaymentsToAcquireMarketableSecurities'],
            'PaymentsForProceedsFromOtherInvestingActivities': ['PaymentsForProceedsFromOtherInvestingActivities'],
            'NetCashProvidedByUsedInInvestingActivities': [
                'NetCashProvidedByUsedInInvestingActivities',
                'NetCashProvidedByUsedInInvestingActivitiesContinuingOperations'
            ],
            
            'PaymentsOfDividendsCommonStock': ['PaymentsOfDividendsCommonStock', 'PaymentsOfDividends'],
            'ProceedsFromIssuanceOrSaleOfEquity': [
                'ProceedsFromIssuanceOrSaleOfEquity', 'ProceedsFromIssuanceOfCommonStock',
                'ProceedsFromShareBasedCompensationArrangements'
            ],
            'PaymentsForRepurchaseOfCommonStock': [
                'PaymentsForRepurchaseOfCommonStock', 'PaymentsForRepurchaseOfEquity'
            ],
            'PaymentsForExciseTaxOnPurchaseOfCommonStock': [
                'PaymentsForExciseTaxOnPurchaseOfCommonStock', 'ExciseTaxOnShareRepurchasesPaid'
            ],
            'ProceedsFromDebtMaturingInMoreThanThreeMonths': [
                'ProceedsFromDebtMaturingInMoreThanThreeMonths', 'ProceedsFromIssuanceOfLongTermDebt'
            ],
            'RepaymentsOfDebtMaturingInMoreThanThreeMonths': [
                'RepaymentsOfDebtMaturingInMoreThanThreeMonths', 'RepaymentsOfLongTermDebt'
            ],
            'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess': [
                'ProceedsFromRepaymentsOfShortTermDebtMaturingInThreeMonthsOrLess',
                'ProceedsFromRepaymentsOfShortTermDebt', 'NetBorrowingsUnderLineOfCredit'
            ],
            'NetCashProvidedByUsedInFinancingActivities': [
                'NetCashProvidedByUsedInFinancingActivities',
                'NetCashProvidedByUsedInFinancingActivitiesContinuingOperations'
            ],
            
            'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents': [
                'EffectOfExchangeRateOnCashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents',
                'EffectOfExchangeRateOnCashAndCashEquivalents'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseIncludingExchangeRateEffect',
                'CashAndCashEquivalentsPeriodIncreaseDecrease'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_Beginning': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsAtCarryingValueBeginningBalance',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsBeginningBalance',
                'CashAndCashEquivalentsAtCarryingValueBeginningBalance'
            ],
            'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents_End': [
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsAtCarryingValue',
                'CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents',
                'CashAndCashEquivalentsAtCarryingValue'
            ],
        }

    # -------------------------------------------------------------------------
    # EDGAR Filing Retrieval (Robust Instance Discovery)
    # -------------------------------------------------------------------------
    def get_all_filings(self, start_year=2020):
        """Get all 10-Q and 10-K filings from start_year to present"""
        url = f"{self.base_url}/submissions/CIK{self.cik}.json"
        logger.info(f"Fetching all filings since {start_year} for {self.company_name}")
        
        try:
            response = requests.get(url, headers=self.headers)
            response.raise_for_status()
            time.sleep(0.2)  # SEC rate limiting
            
            data = response.json()
            recent = data['filings']['recent']
            
            filings = []
            for i in range(len(recent['form'])):
                filing_date = recent['filingDate'][i]
                filing_year = int(filing_date.split('-')[0])
                
                if filing_year >= start_year:
                    form = recent['form'][i]
                    if form in ['10-Q', '10-K']:
                        filings.append({
                            'accession': recent['accessionNumber'][i],
                            'filing_date': filing_date,
                            'report_date': recent['reportDate'][i],
                            'form': form,
                            'primary_document': recent['primaryDocument'][i]
                        })
            
            filings.sort(key=lambda x: x['report_date'])
            logger.info(f"Found {len(filings)} filings since {start_year}")
            return filings
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Error fetching submissions: {e}")
            raise

    def _filing_base_dir(self, accession: str) -> str:
        """Return folder URL for a filing (no trailing slash)"""
        accession_no_dash = accession.replace('-', '')
        return f"{self.sec_archives}/{self.cik_int}/{accession_no_dash}"

    def get_filing_items(self, accession: str):
        """
        Return directory items using index.json or HTML directory fallback.
        This prevents 404 errors from hardcoded filename assumptions.
        """
        base_dir = self._filing_base_dir(accession)
        
        # Try index.json first (preferred method)
        idx_url = f"{base_dir}/index.json"
        try:
            r = requests.get(idx_url, headers=self.headers)
            if r.status_code == 200:
                time.sleep(0.2)
                j = r.json()
                items = j.get('directory', {}).get('item', [])
                # Normalize to list of dicts with at least "name"
                return [{'name': it.get('name', ''), 'type': it.get('type', '')} for it in items]
        except Exception:
            pass  # Fall through to HTML parsing
        
        # Fallback: parse HTML directory listing
        try:
            r = requests.get(base_dir, headers=self.headers)
            r.raise_for_status()
            time.sleep(0.2)
            html = r.text
            
            # Extract hrefs from HTML
            hrefs = re.findall(r'href="([^"]+)"', html, flags=re.IGNORECASE)
            # Keep only same-folder filenames (no parent links or external URLs)
            names = [h for h in hrefs 
                    if not h.startswith('http') 
                    and not h.startswith('?') 
                    and '/' not in h.strip('/')]
            names = list(dict.fromkeys(names))  # Deduplicate while preserving order
            
            return [{'name': n, 'type': ''} for n in names]
            
        except Exception as e:
            logger.error(f"Unable to list directory for accession {accession}: {e}")
            return []

    def pick_instance_from_items(self, items: list) -> str:
        """
        Choose the XBRL instance document from directory items.
        Uses heuristics to identify the correct .xml file.
        """
        names = [i.get('name') for i in items if i.get('name')]
        if not names:
            return None
        
        # Get all XML files
        xmls = [n for n in names if n.lower().endswith('.xml')]
        
        # Exclude known non-instance files
        EXCLUDE_SUBSTR = ['_cal.xml', '_def.xml', '_lab.xml', '_pre.xml',
                         '.xsd', 'filingsummary', 'metalink', 'schema']
        xmls = [n for n in xmls if all(ex not in n.lower() for ex in EXCLUDE_SUBSTR)]
        
        # Prefer files matching common instance patterns
        preferences = [
            '_htm.xml',           # Common iXBRL format
            f"{self.ticker}-",    # Contains company ticker
        ]
        
        for pref in preferences:
            filtered = [n for n in xmls if pref in n.lower()]
            if filtered:
                # If multiple matches, pick shortest name (usually the main instance)
                return sorted(filtered, key=len)[0]
        
        # Otherwise, pick first remaining XML (stable order from index)
        if xmls:
            return sorted(xmls, key=len)[0]
        
        return None

    def download_file(self, url: str) -> bytes:
        """Download a file from SEC EDGAR with error handling"""
        r = requests.get(url, headers=self.headers)
        r.raise_for_status()
        time.sleep(0.2)  # SEC rate limiting
        return r.content

    # -------------------------------------------------------------------------
    # XBRL Parsing
    # -------------------------------------------------------------------------
    def parse_context_elements(self, root):
        """Parse XBRL context elements to understand segments and time periods"""
        contexts = {}
        
        for context in root.findall('.//xbrli:context', self.namespaces):
            context_id = context.get('id')
            period = context.find('xbrli:period', self.namespaces)
            
            instant = period.find('xbrli:instant', self.namespaces)
            start = period.find('xbrli:startDate', self.namespaces)
            end = period.find('xbrli:endDate', self.namespaces)
            
            context_info = {
                'id': context_id,
                'segments': {},
                'instant': instant.text if instant is not None else None,
                'start': start.text if start is not None else None,
                'end': end.text if end is not None else None,
            }
            
            # Extract segment/dimension information
            entity = context.find('xbrli:entity', self.namespaces)
            if entity is not None:
                segment = entity.find('xbrli:segment', self.namespaces)
                if segment is not None:
                    for member in segment.findall('.//xbrldi:explicitMember', self.namespaces):
                        dimension = member.get('dimension')
                        member_value = member.text
                        # Strip namespace prefix if present
                        if ':' in member_value:
                            member_value = member_value.split(':')[1]
                        context_info['segments'][dimension] = member_value
            
            contexts[context_id] = context_info
        
        return contexts

    def extract_facts_from_xbrl(self, xml_content: bytes):
        """Extract all facts from XBRL instance document"""
        root = ET.fromstring(xml_content)
        
        # Update namespaces from document
        for prefix, uri in root.attrib.items():
            if prefix.startswith('{http://www.w3.org/2000/xmlns/}'):
                ns_prefix = prefix.split('}')[1]
                self.namespaces[ns_prefix] = uri
        
        contexts = self.parse_context_elements(root)
        facts = []
        
        # Iterate through all elements looking for facts
        for elem in root.iter():
            tag_name = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
            context_ref = elem.get('contextRef')
            unit_ref = elem.get('unitRef')
            decimals = elem.get('decimals')
            
            if context_ref in contexts and elem.text:
                context = contexts[context_ref]
                
                # Default to consolidated unless a segment is present
                segment_name = "Consolidated"
                segment_dimension = None
                business_segment = None
                if context['segments']:
                    for dim, member in context['segments'].items():
                        # Capture business segment axis separately
                        if 'StatementBusinessSegmentsAxis' in dim:
                            business_segment = member
                        if segment_name == "Consolidated":
                            segment_name = member
                            segment_dimension = dim
                    # If we found a business segment, prefer it as the primary segment
                    # for multi-dimensional contexts (e.g. segment + consolidation axis)
                    if business_segment and segment_name != business_segment:
                        # Keep business_segment separate; segment stays as-is for backward compat
                        pass
                
                # Try to convert to numeric value
                try:
                    value = float(elem.text)
                except (ValueError, TypeError):
                    continue
                
                facts.append({
                    'tag': tag_name,
                    'value': value,
                    'context_id': context_ref,
                    'segment': segment_name,
                    'dimension': segment_dimension,
                    'business_segment': business_segment,
                    'start_date': context['start'],
                    'end_date': context['end'],
                    'instant_date': context['instant'],
                    'decimals': decimals,
                    'unit': unit_ref
                })
        
        return facts

    def process_filing(self, filing: dict):
        """
        Process a single filing - discovers and downloads the XBRL instance.
        Uses robust discovery to avoid 404 errors from filename assumptions.
        """
        try:
            base_dir = self._filing_base_dir(filing['accession'])
            logger.info(f"Processing {filing['form']} from {filing['report_date']}")
            
            # Discover the actual XBRL instance filename
            items = self.get_filing_items(filing['accession'])
            instance_name = self.pick_instance_from_items(items)
            
            # Fallback: scan primary document HTML for .xml instance link
            if not instance_name and filing.get('primary_document'):
                primary_url = f"{base_dir}/{filing['primary_document']}"
                try:
                    html = self.download_file(primary_url).decode('utf-8', errors='ignore')
                    # Look for .xml links, exclude non-instance patterns
                    links = re.findall(r'href="([^"]+\.xml)"', html, flags=re.IGNORECASE)
                    links = [l for l in links if not any(s in l.lower() for s in
                                                         ['_cal.xml', '_def.xml', '_lab.xml', '_pre.xml',
                                                          '.xsd', 'filingsummary', 'metalink', 'schema'])]
                    if links:
                        # Normalize to filename only
                        instance_name = links[0].split('/')[-1]
                except Exception:
                    pass
            
            if not instance_name:
                logger.warning(f"Could not locate XBRL instance for accession {filing['accession']} - skipping")
                return []
            
            # Download and parse the instance
            instance_url = f"{base_dir}/{instance_name}"
            xml_content = self.download_file(instance_url)
            facts = self.extract_facts_from_xbrl(xml_content)
            
            # Annotate facts with filing metadata
            for fact in facts:
                fact['accession'] = filing['accession']
                fact['filing_date'] = filing['filing_date']
                fact['report_date'] = filing['report_date']
                fact['form'] = filing['form']
            
            logger.info(f"  Extracted {len(facts)} facts from {instance_name}")
            return facts
            
        except requests.HTTPError as e:
            logger.error(f"HTTP error for filing {filing.get('accession')}: {e}")
            return []
        except Exception as e:
            logger.error(f"Error processing filing: {e}")
            return []

    def extract_all_data(self, start_year=2020):
        """Extract all financial data from filings"""
        filings = self.get_all_filings(start_year=start_year)
        all_facts = []
        
        for i, filing in enumerate(filings, 1):
            logger.info(f"\n[{i}/{len(filings)}] " + "=" * 50)
            facts = self.process_filing(filing)
            all_facts.extend(facts)
        
        # Convert to DataFrame and process dates
        df = pd.DataFrame(all_facts)
        if not df.empty:
            for date_col in ['start_date', 'end_date', 'instant_date', 'filing_date', 'report_date']:
                if date_col in df.columns:
                    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            
            sort_col = 'end_date' if 'end_date' in df.columns else 'instant_date'
            df = df.sort_values([sort_col, 'tag', 'segment'], ascending=[True, True, True])

            before = len(df)
            df = df.drop_duplicates()
            if len(df) < before:
                logger.info(f"Removed {before - len(df)} exact duplicate facts")

            logger.info(f"\nTotal facts extracted: {len(df)}")
        
        return df

    # -------------------------------------------------------------------------
    # Q4 Calculation Logic (YTD-aware)
    # -------------------------------------------------------------------------
    def _is_calendar_ytd(self, row):
        """Determine if a period is YTD for calendar fiscal year"""
        s = row.get('start_date')
        e = row.get('end_date')
        if pd.isna(s) or pd.isna(e):
            return False
        # Calendar FY: starts Jan 1 and ends in same year
        return s.month == 1 and s.day == 1 and s.year == e.year

    def calculate_q4_data(self, df):
        """
        Calculate Q4 data using YTD-aware logic:
        - Balance sheet: Copy 10-K instant values
        - Flow statements: Annual - Q3 (YTD) or Annual - sum(Q1+Q2+Q3)
        """
        logger.info("\n" + "=" * 60)
        logger.info("Calculating Q4 data")
        logger.info("=" * 60)
        
        if df.empty:
            return df
        
        quarterly_df = df[df['form'] == '10-Q'].copy()
        annual_df = df[df['form'] == '10-K'].copy()
        
        q4_records = []
        unique_tags = df['tag'].dropna().unique()
        unique_segments = df['segment'].dropna().unique().tolist()
        
        if 'Consolidated' not in unique_segments:
            unique_segments.append('Consolidated')
        
        for tag in unique_tags:
            for segment in unique_segments:
                # Get annual data for this tag/segment combination
                annual_subset = annual_df[
                    (annual_df['tag'] == tag) &
                    ((annual_df['segment'] == segment) |
                     (segment == 'Consolidated' and annual_df['segment'].isna()))
                ].copy()
                
                for _, annual_row in annual_subset.iterrows():
                    # Get the business_segment of this annual row (if any)
                    biz_seg = annual_row.get('business_segment')
                    
                    # Balance sheet items (instant dates)
                    is_balance_sheet = pd.notna(annual_row.get('instant_date'))
                    
                    if is_balance_sheet:
                        fiscal_year_end = annual_row['instant_date']
                        if pd.isna(fiscal_year_end):
                            continue
                        
                        q4_records.append({
                            'tag': tag,
                            'value': annual_row['value'],
                            'segment': segment,
                            'business_segment': annual_row.get('business_segment'),
                            'start_date': None,
                            'end_date': None,
                            'instant_date': fiscal_year_end,
                            'report_date': fiscal_year_end,
                            'form': '10-K (Q4)',
                            'accession': annual_row['accession'],
                            'filing_date': annual_row['filing_date'],
                            'context_id': f"Q4_{fiscal_year_end.year}_{segment}",
                            'dimension': annual_row.get('dimension'),
                            'decimals': annual_row.get('decimals'),
                            'unit': annual_row.get('unit'),
                        })
                        continue
                    
                    # Flow statement items (period data)
                    fiscal_year_end = annual_row['end_date']
                    if pd.isna(fiscal_year_end):
                        continue
                    
                    fiscal_year = fiscal_year_end.year
                    
                    # Get quarterly data for this fiscal year
                    q_subset = quarterly_df[
                        (quarterly_df['tag'] == tag) &
                        ((quarterly_df['segment'] == segment) |
                         (segment == 'Consolidated' and quarterly_df['segment'].isna())) &
                        (quarterly_df['end_date'] > pd.Timestamp(year=fiscal_year - 1, month=12, day=31)) &
                        (quarterly_df['end_date'] <= fiscal_year_end)
                    ].copy()
                    
                    # If this annual row has a business_segment, filter quarterly data to match
                    if pd.notna(biz_seg) and biz_seg:
                        q_biz = q_subset[q_subset['business_segment'] == biz_seg]
                        if not q_biz.empty:
                            q_subset = q_biz
                        # else fall back to unfiltered (for backward compatibility)
                    
                    if q_subset.empty:
                        continue
                    
                    # Detect YTD reporting
                    q_subset['is_ytd'] = q_subset.apply(self._is_calendar_ytd, axis=1)
                    
                    annual_total = annual_row['value']
                    q3_end = q_subset['end_date'].max()
                    
                    # Calculate Q4 value
                    if q_subset['is_ytd'].any():
                        # YTD reporting: Q4 = Annual - Latest YTD (typically Q3)
                        latest_ytd = q_subset.sort_values('end_date').iloc[-1]
                        q4_value = annual_total - latest_ytd['value']
                    else:
                        # Discrete quarters: Q4 = Annual - sum(available quarters)
                        sum_quarters = q_subset['value'].sum()
                        q4_value = annual_total - sum_quarters
                        if len(q_subset) != 3:
                            logger.warning(
                                f"Discrete quarterly data incomplete for {tag} ({segment}) FY{fiscal_year}: "
                                f"have {len(q_subset)} quarters; Q4 computed as Annual - sum(available)"
                            )
                    
                    q4_records.append({
                        'tag': tag,
                        'value': q4_value,
                        'segment': segment,
                        'business_segment': annual_row.get('business_segment'),
                        'start_date': q3_end + pd.Timedelta(days=1) if pd.notna(q3_end) else None,
                        'end_date': fiscal_year_end,
                        'instant_date': None,
                        'report_date': fiscal_year_end,
                        'form': '10-Q (Q4 Calculated)',
                        'accession': annual_row['accession'],
                        'filing_date': annual_row['filing_date'],
                        'context_id': f"Q4_{fiscal_year}_{segment}",
                        'dimension': annual_row.get('dimension'),
                        'decimals': annual_row.get('decimals'),
                        'unit': annual_row.get('unit'),
                    })
        
        if q4_records:
            q4_df = pd.DataFrame(q4_records)
            combined = pd.concat([df, q4_df], ignore_index=True)
            if 'end_date' in combined.columns:
                combined = combined.sort_values(['end_date', 'instant_date', 'tag', 'segment'])
            logger.info(f"Added {len(q4_records)} Q4 records")
            return combined
        
        return df

    def _normalize_quarters_to_discrete(self, df_quarters: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize YTD quarterly cash flow to discrete quarters.
        Q1 = reported, Q2 = Q2-Q1, Q3 = Q3-Q2
        (Q4 is handled separately by calculate_q4_data)
        """
        if df_quarters.empty:
            return df_quarters
        
        q = df_quarters.copy()
        q = q[q['form'].str.contains('10-Q', na=False)]
        q = q[q['end_date'].notna()].copy()
        q['segment'] = q['segment'].fillna('Consolidated')
        q['fiscal_year'] = q['end_date'].dt.year
        q['is_ytd'] = q.apply(self._is_calendar_ytd, axis=1)
        
        # Include business_segment in groupby to avoid mixing segment data during diff
        if 'business_segment' not in q.columns:
            q['business_segment'] = None
        q['_biz_seg_key'] = q['business_segment'].fillna('__none__')
        
        q = q.sort_values(['tag', 'segment', '_biz_seg_key', 'fiscal_year', 'end_date'])
        
        # Normalize YTD to discrete per group (avoids FutureWarning from groupby.apply)
        result_parts = []
        for _, g in q.groupby(['tag', 'segment', '_biz_seg_key', 'fiscal_year'], group_keys=False):
            if g['is_ytd'].any():
                g = g.sort_values('end_date').copy()
                g['value'] = g['value'].diff().fillna(g['value'])
            result_parts.append(g)
        q = pd.concat(result_parts, ignore_index=True) if result_parts else q
        q = q.drop(columns=['_biz_seg_key'])
        return q

    # -------------------------------------------------------------------------
    # Pivot Table Creation
    # -------------------------------------------------------------------------
    def create_statement_pivot(self, df, statement_type):
        """Create pivot table for financial statement with proper line item order"""
        if df.empty:
            return pd.DataFrame()
        
        statement_items = self._get_statement_items(statement_type)
        
        # Determine date column based on statement type
        if statement_type == 'balance':
            date_col = 'instant_date'
            df_filtered = df[df['instant_date'].notna()].copy()
        else:
            date_col = 'end_date'
            df_filtered = df[df['end_date'].notna()].copy()
        
        if df_filtered.empty:
            return pd.DataFrame()
        
        # Filter to quarterly data only (includes "10-Q (Q4 Calculated)")
        df_filtered = df_filtered[df_filtered['form'].str.contains('10-Q', na=False)].copy()
        
        # For cash flow: normalize Q1/Q2/Q3 to discrete, keep Q4 Calculated as-is
        if statement_type == 'cashflow':
            is_q4_calc = df_filtered['form'].str.contains('Q4 Calculated', na=False)
            q4_calc = df_filtered[is_q4_calc].copy()
            q10 = df_filtered[~is_q4_calc].copy()
            q10 = self._normalize_quarters_to_discrete(q10)
            df_filtered = pd.concat([q10, q4_calc], ignore_index=True)
        
        df_filtered['segment'] = df_filtered['segment'].fillna('Consolidated')
        
        # Get candidate tag maps
        income_map = self._get_income_tag_candidates() if statement_type == 'income' else {}
        balance_map = self._get_balance_tag_candidates() if statement_type == 'balance' else {}
        cash_map = self._get_cashflow_tag_candidates() if statement_type == 'cashflow' else {}
        
        # Segment suffix mapping
        segment_map = {
            'FinancialProducts': 'FinancialProductsMember',
            'MET': 'MachineryEnergyTransportationMember',
            'EXFP': 'AllOtherExcludingFinancialProductsMember',
            'Total': 'Consolidated',
        }
        
        pivot_data = []
        
        for tag_key, label in statement_items.items():
            # Handle blank label rows (section headers)
            if tag_key == '':
                pivot_data.append({'Line_Item': label})
                continue
            
            # Extract base tag and optional segment suffix
            parts = tag_key.split('_', 1)
            base_key = parts[0]
            segment_suffix = parts[1] if len(parts) > 1 else None
            
            # Determine candidate tags based on statement type
            if statement_type == 'income':
                candidate_tags = income_map.get(base_key, [base_key])
            elif statement_type == 'balance':
                candidate_tags = balance_map.get(base_key, [base_key if segment_suffix else tag_key])
            elif statement_type == 'cashflow':
                candidate_tags = cash_map.get(base_key, [base_key if segment_suffix else tag_key])
            else:
                candidate_tags = [tag_key]
            
            selected_subset = pd.DataFrame()
            
            # Find data matching the tag and segment
            if segment_suffix:
                target_segment = segment_map.get(segment_suffix, segment_suffix)
                # Try exact match first
                for cand in candidate_tags:
                    sub = df_filtered[
                        (df_filtered['tag'] == cand) & 
                        (df_filtered['segment'] == target_segment)
                    ]
                    if not sub.empty:
                        selected_subset = sub
                        break
                
                # Fallback: case-insensitive contains
                if selected_subset.empty:
                    for cand in candidate_tags:
                        sub = df_filtered[
                            (df_filtered['tag'] == cand) &
                            (df_filtered['segment'].str.contains(target_segment, case=False, na=False))
                        ]
                        if not sub.empty:
                            selected_subset = sub
                            break
            else:
                # Consolidated data (no segment)
                for cand in candidate_tags:
                    sub = df_filtered[
                        (df_filtered['tag'] == cand) &
                        (df_filtered['segment'].isin(['Consolidated', '']))
                    ]
                    if not sub.empty:
                        selected_subset = sub
                        break
            
            # Create pivot row if data found
            if not selected_subset.empty:
                pv = selected_subset.pivot_table(
                    index='tag', 
                    columns=date_col, 
                    values='value', 
                    aggfunc='first'
                )
                if not pv.empty:
                    row = {'Line_Item': label}
                    row.update(pv.iloc[0].to_dict())
                    pivot_data.append(row)
        
        if not pivot_data:
            return pd.DataFrame()
        
        result_df = pd.DataFrame(pivot_data)
        
        # Sort columns by date (most recent first), keep Line_Item first
        date_cols = [c for c in result_df.columns if c != 'Line_Item']
        date_cols_sorted = sorted(date_cols)
        result_df = result_df[['Line_Item'] + date_cols_sorted]
        
        # Format datetime column names as YYYY-MM-DD
        formatted = []
        for col in result_df.columns:
            if isinstance(col, pd.Timestamp):
                formatted.append(col.strftime('%Y-%m-%d'))
            else:
                formatted.append(col)
        result_df.columns = formatted
        
        return result_df

    def _get_segment_tag_candidates(self):
        """Candidate tags for segment items (handles XBRL tag variations across filings)"""
        return {
            'Revenues': ['Revenues', 'RevenueFromContractWithCustomerExcludingAssessedTax', 'SalesRevenueNet'],
            'CostOfRevenue': ['CostOfRevenue', 'CostOfGoodsAndServicesSold', 'CostOfSales'],
            'SellingGeneralAndAdministrativeResearchAndDevelopment': [
                'SellingGeneralAndAdministrativeResearchAndDevelopment'],
            'SegmentReportingOtherItemAmount': ['SegmentReportingOtherItemAmount'],
            'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments': [
                'IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments',
                'IncomeLossBeforeTaxesAfterRestructuringCosts'],
            'Assets': ['Assets'],
            'DepreciationDepletionAndAmortization': ['DepreciationDepletionAndAmortization'],
            'SegmentReportingInformationExpenditureForAdditionsLongLivedAssets': [
                'SegmentReportingInformationExpenditureForAdditionsLongLivedAssets',
                'SegmentExpenditureAdditionToLongLivedAssets'],
        }

    def _extract_discrete_quarters(self, flow_df):
        """Extract discrete quarterly values from flow data that may contain both
        discrete (3-month) and YTD (6-month, 9-month) values.
        
        XBRL 10-Q filings report BOTH discrete and YTD figures for Q2/Q3.
        Applying diff() to mixed data produces garbage. Instead:
        1. Prefer discrete values (period ≤ ~100 days) when available
        2. Fall back to YTD normalization only for periods with no discrete data
        """
        if flow_df.empty:
            return flow_df
        
        q = flow_df.copy()
        q = q[q['start_date'].notna() & q['end_date'].notna()].copy()
        
        if q.empty:
            return flow_df
        
        # Compute period length in days
        q['_period_days'] = (q['end_date'] - q['start_date']).dt.days
        
        # Discrete quarters: ~90 days (allow up to 100 for slight variations)
        # YTD periods: H1 ~180 days, 9M ~270 days
        discrete_mask = q['_period_days'] <= 100
        discrete_df = q[discrete_mask].copy()
        ytd_df = q[~discrete_mask].copy()
        
        if not discrete_df.empty:
            # We have discrete values — use them
            # Check if any (tag, end_date) combos are missing discrete but have YTD
            discrete_keys = set(zip(discrete_df['tag'], discrete_df['end_date']))
            
            if not ytd_df.empty:
                # Keep only YTD rows for periods NOT covered by discrete data
                ytd_needed = ytd_df[~ytd_df.apply(
                    lambda r: (r['tag'], r['end_date']) in discrete_keys, axis=1
                )]
                if not ytd_needed.empty:
                    # Normalize YTD-only periods to discrete
                    ytd_needed = self._normalize_quarters_to_discrete(ytd_needed)
                    result = pd.concat([discrete_df, ytd_needed], ignore_index=True)
                else:
                    result = discrete_df
            else:
                result = discrete_df
        else:
            # Only YTD data — normalize via diff
            result = self._normalize_quarters_to_discrete(ytd_df)
        
        result = result.drop(columns=['_period_days'], errors='ignore')
        return result

    def create_segment_pivot(self, df, segment_member):
        """Create pivot table for a specific business segment.
        
        segment_member: e.g. 'ConstructionIndustriesMember'
        
        Data comes from two sources:
        1. Direct tags: segment field == segment_member (via StatementBusinessSegmentsAxis)
        2. Multi-dimensional tags: business_segment == segment_member (COGS, SGA, etc.)
        """
        if df.empty:
            return pd.DataFrame()
        
        segment_items = self._get_segment_items()
        tag_candidates = self._get_segment_tag_candidates()
        
        # Balance-sheet-like items (point-in-time) vs flow items
        instant_tags = {'Assets'}
        
        # Merge: rows where segment==segment_member OR business_segment==segment_member
        mask_direct = df['segment'] == segment_member
        mask_biz = df['business_segment'] == segment_member
        seg_df = df[mask_direct | mask_biz].copy()
        
        if seg_df.empty:
            logger.warning(f"No data found for segment: {segment_member}")
            return pd.DataFrame()
        
        seg_df = seg_df.drop_duplicates()
        # Resolve candidate tags: for each segment item, find the best matching
        # XBRL tag and remap to the canonical key
        for canonical_tag, candidates in tag_candidates.items():
            if len(candidates) <= 1:
                continue
            # Check which candidate tags have data
            for alt_tag in candidates[1:]:  # Skip first (it IS the canonical)
                alt_mask = seg_df['tag'] == alt_tag
                if alt_mask.any():
                    # Check if canonical already has data for the same periods
                    canon_mask = seg_df['tag'] == canonical_tag
                    if not canon_mask.any():
                        # No canonical data — remap alt to canonical
                        seg_df.loc[alt_mask, 'tag'] = canonical_tag
                    else:
                        # Both exist — remap alt rows for periods where canonical is missing
                        canon_dates = set(seg_df.loc[canon_mask, 'end_date'].dropna())
                        alt_new = alt_mask & ~seg_df['end_date'].isin(canon_dates)
                        seg_df.loc[alt_new, 'tag'] = canonical_tag
        
        # Filter to quarterly data (10-Q and Q4 calculated)
        seg_df = seg_df[seg_df['form'].str.contains('10-Q', na=False)].copy()
        
        if seg_df.empty:
            return pd.DataFrame()
        
        # Add source priority: direct segment tags (priority 0) beat
        # multi-dimensional OperatingSegmentsMember tags (priority 1)
        seg_df['_src_priority'] = 1  # default: multi-dim
        seg_df.loc[seg_df['segment'] == segment_member, '_src_priority'] = 0  # direct
        seg_df = seg_df.sort_values('_src_priority')
        
        # Unify the segment field: all rows in this pivot belong to the same 
        # business segment, so normalize segment to avoid split groups
        seg_df['segment'] = segment_member
        
        # Deduplicate: if both direct and multi-dimensional sources have same 
        # tag + period, keep the direct one (lower priority number = first)
        seg_df = seg_df.drop_duplicates(
            subset=['tag', 'start_date', 'end_date', 'instant_date', 'form'], 
            keep='first'
        )
        # Also catch exact duplicate facts (same tag+value+dates from different contexts)
        seg_df = seg_df.drop_duplicates(
        subset=['tag', 'value', 'start_date', 'end_date', 'instant_date'], 
        keep='first'
        )
        seg_df = seg_df.drop(columns=['_src_priority'])
        
        # --- Handle flow items: extract discrete quarters ---
        is_instant_tag = seg_df['tag'].isin(instant_tags)
        instant_df = seg_df[is_instant_tag].copy()
        flow_df = seg_df[~is_instant_tag].copy()
        
        if not flow_df.empty:
            is_q4_calc = flow_df['form'].str.contains('Q4 Calculated', na=False)
            q4_calc = flow_df[is_q4_calc].copy()
            q10 = flow_df[~is_q4_calc].copy()
            
            # Use discrete-aware extraction instead of blind YTD normalization
            if not q10.empty:
                q10 = self._extract_discrete_quarters(q10)
            flow_df = pd.concat([q10, q4_calc], ignore_index=True)
        
        # Recombine
        seg_df = pd.concat([instant_df, flow_df], ignore_index=True)
        
        # --- Build pivot rows ---
        pivot_data = []
        
        for tag_key, label in segment_items.items():
            if tag_key == '':
                pivot_data.append({'Line_Item': label})
                continue
            
            is_instant = tag_key in instant_tags
            date_col = 'instant_date' if is_instant else 'end_date'
            
            sub = seg_df[seg_df['tag'] == tag_key].copy()
            
            if sub.empty:
                continue
            
            # Deduplicate: keep one value per date
            sub = sub.dropna(subset=[date_col])
            sub = sub.drop_duplicates(subset=['tag', date_col], keep='first')
            
            pv = sub.pivot_table(
                index='tag',
                columns=date_col,
                values='value',
                aggfunc='first'
            )
            
            if not pv.empty:
                row = {'Line_Item': label}
                row.update(pv.iloc[0].to_dict())
                pivot_data.append(row)
        
        if not pivot_data:
            return pd.DataFrame()
        
        result_df = pd.DataFrame(pivot_data)
        
        # Sort columns by date (Oldest first)
        date_cols = [c for c in result_df.columns if c != 'Line_Item']
        date_cols_sorted = sorted(date_cols)
        result_df = result_df[['Line_Item'] + date_cols_sorted]
        
        # Format datetime column names
        formatted = []
        for col in result_df.columns:
            if isinstance(col, pd.Timestamp):
                formatted.append(col.strftime('%Y-%m-%d'))
            else:
                formatted.append(col)
        result_df.columns = formatted
        
        return result_df

    # -------------------------------------------------------------------------
    # Excel Formatting
    # -------------------------------------------------------------------------
    def _get_quarter_from_date(self, date_str):
        """Determine quarter label from date string"""
        try:
            date_obj = pd.to_datetime(date_str) if isinstance(date_str, str) else date_str
            m = date_obj.month
            if m in [1, 2, 3]:
                return 'Q1'
            elif m in [4, 5, 6]:
                return 'Q2'
            elif m in [7, 8, 9]:
                return 'Q3'
            elif m in [10, 11, 12]:
                return 'Q4'
            else:
                return ''
        except Exception:
            return ''

    def format_excel_sheet(self, writer, sheet_name, df):
        """Apply professional formatting to Excel sheet with quarter labels"""
        ws = writer.sheets[sheet_name]
        
        # Insert row at top for quarter labels
        ws.insert_rows(1)
        
        # Add quarter labels for each date column
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name != 'Line_Item' and isinstance(col_name, str):
                quarter = self._get_quarter_from_date(col_name)
                cell = ws.cell(row=1, column=col_idx)
                cell.value = quarter
        
        # Format quarter row (row 1)
        quarter_format = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        quarter_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[1]:
            cell.fill = quarter_format
            cell.font = quarter_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format date header row (row 2)
        header_format = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[2]:
            cell.fill = header_format
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for column in ws.columns:
            max_len = 0
            col_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    txt = str(cell.value) if cell.value is not None else ""
                    if len(txt) > max_len:
                        max_len = len(txt)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        # Format number cells (starting from row 3)
        accounting_format = '#,##0'
        for row in ws.iter_rows(min_row=3):
            for cell in row[1:]:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = accounting_format
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Freeze panes to keep headers visible
        ws.freeze_panes = 'B3'

    # -------------------------------------------------------------------------
    # Main Export Function
    # -------------------------------------------------------------------------
    def export_to_excel(self, output_filename, start_year=2020):
        """Extract all data and export to professionally formatted Excel file"""
        logger.info("=" * 60)
        logger.info(f"Starting comprehensive extraction for {self.company_name}")
        logger.info(f"Data range: {start_year} - Present")
        logger.info("=" * 60)
        
        # Extract all data
        df = self.extract_all_data(start_year)
        if df.empty:
            logger.warning("No data extracted!")
            return None
        
        # Calculate Q4 data
        df = self.calculate_q4_data(df)
        
        # Create Excel file with multiple sheets
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Raw data sheet
            df.to_excel(writer, sheet_name='All Data - Raw', index=False)
            self.format_excel_sheet(writer, 'All Data - Raw', df)
            
            # Income Statement
            logger.info("Creating Income Statement")
            income_pivot = self.create_statement_pivot(df, 'income')
            if not income_pivot.empty:
                income_pivot.to_excel(writer, sheet_name='Income Statement - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Income Statement - Quarterly', income_pivot)
            
            # Balance Sheet
            logger.info("Creating Balance Sheet")
            balance_pivot = self.create_statement_pivot(df, 'balance')
            if not balance_pivot.empty:
                balance_pivot.to_excel(writer, sheet_name='Balance Sheet - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Balance Sheet - Quarterly', balance_pivot)
            
            # Cash Flow Statement
            logger.info("Creating Cash Flow Statement")
            cashflow_pivot = self.create_statement_pivot(df, 'cashflow')
            if not cashflow_pivot.empty:
                cashflow_pivot.to_excel(writer, sheet_name='Cash Flow - Quarterly', index=False)
                self.format_excel_sheet(writer, 'Cash Flow - Quarterly', cashflow_pivot)
            
            # Segment Sheets
            segment_configs = [
                ('ConstructionIndustriesMember', 'Construction Industries'),
                ('ResourceIndustriesMember', 'Resource Industries'),
                ('EnergyandTransportationMember', 'Energy & Transportation'),
                ('FinancialProductsSegmentMember', 'Financial Products Segment'),
            ]
            
            for member_name, sheet_name in segment_configs:
                logger.info(f"Creating {sheet_name} segment sheet")
                seg_pivot = self.create_segment_pivot(df, member_name)
                # Fallback: Financial Products may use alternate member name
                if seg_pivot.empty and 'FinancialProducts' in member_name:
                    alt_name = 'FinancialProductsMember' if member_name == 'FinancialProductsSegmentMember' else 'FinancialProductsSegmentMember'
                    logger.info(f"  Trying alternate member: {alt_name}")
                    seg_pivot = self.create_segment_pivot(df, alt_name)
                if not seg_pivot.empty:
                    seg_pivot.to_excel(writer, sheet_name=sheet_name, index=False)
                    self.format_excel_sheet(writer, sheet_name, seg_pivot)
                else:
                    logger.warning(f"No data for segment: {sheet_name}")
        
        logger.info("=" * 60)
        logger.info(f"Export complete! File saved: {output_filename}")
        logger.info("=" * 60)
        
        return output_filename


# =============================================================================
# Main Execution
# =============================================================================
def main():
    """Main execution function"""
    # Configuration
    YOUR_EMAIL = "brayden.joyce@doosan.com"  # Required by SEC for identification
    CIK = "0000018230"  # Caterpillar Inc.
    COMPANY_NAME = "Caterpillar Inc."
    TICKER = "cat"
    START_YEAR = 2020
    
    # Create extractor instance
    extractor = ComprehensiveXBRLExtractor(
        email=YOUR_EMAIL,
        cik=CIK,
        company_name=COMPANY_NAME,
        ticker=TICKER
    )
    
    # Extract and export data
    output_file = extractor.export_to_excel(
        output_filename='caterpillar_financials.xlsx',
        start_year=START_YEAR
    )
    
    if output_file:
        print(f"\n{'='*60}")
        print(f"SUCCESS! Complete financial data exported to: {output_file}")
        print(f"\nData range: {START_YEAR} - Present")
        print("\nThe Excel file contains:")
        print("  Income Statement - Quarterly")
        print("  Balance Sheet - Quarterly")
        print("  Cash Flow Statement - Quarterly")
        print("  Construction Industries - Segment")
        print("  Resource Industries - Segment")
        print("  Energy & Transportation - Segment")
        print("  Financial Products - Segment")
        print("  All raw data with segment breakdowns")
        print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
